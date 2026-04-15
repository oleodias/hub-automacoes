from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response, send_file
import os
import sys
import json
from datetime import datetime, time
import subprocess
import requests
import threading
import time as time_module
from werkzeug.utils import secure_filename
from automacoes.clientes.leitor_ficha import processar_ficha_cliente
from automacoes.clientes import cadastro_novo
from automacoes.clientes import cadastro_reativacao
import re
import requests
from bs4 import BeautifulSoup
from flask import jsonify, render_template
import urllib3
import requests as req_ext
import uuid as uuid_lib          # ← NOVO: para gerar o UUID de cada submissão
import banco_cadastros           # ← NOVO: módulo SQLite de cadastros

from dotenv import load_dotenv

load_dotenv() # Carrega as variáveis de ambiente

app = Flask(__name__, template_folder='templates_reestilizados')
# Substitua a chave fixa por esta linha:
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'chave_reserva_caso_falhe')

# Desliga os avisos chatos de segurança quando usamos verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── SISTEMA DE FILA ──────────────────────
_lock_execucao = threading.Lock()
_fila = []
_em_execucao = None

def _gerar_id():
    return f"{int(time_module.time() * 1000)}"

def entrar_na_fila():
    global _fila
    with _lock_execucao:
        meu_id = _gerar_id()
        _fila.append(meu_id)
        return meu_id

def minha_vez(meu_id):
    global _em_execucao, _fila
    with _lock_execucao:
        if _em_execucao is not None:
            return False
        return bool(_fila and _fila[0] == meu_id)

def iniciar_execucao(meu_id):
    global _em_execucao, _fila
    with _lock_execucao:
        _em_execucao = meu_id
        if meu_id in _fila:
            _fila.remove(meu_id)

def finalizar_execucao():
    global _em_execucao
    with _lock_execucao:
        _em_execucao = None

def sair_da_fila(meu_id):
    global _fila
    with _lock_execucao:
        if meu_id in _fila:
            _fila.remove(meu_id)

def posicao_na_fila(meu_id):
    global _fila, _em_execucao
    with _lock_execucao:
        if _em_execucao == meu_id:
            return 0
        if meu_id in _fila:
            return _fila.index(meu_id) + 1
        return -1
# ─────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════════════
# CONSULTA CNPJ.WS — Enriquecimento dos dados da ficha
# ══════════════════════════════════════════════════════════════════════════════
# Esta função é chamada pelo /receber_ficha quando o cliente envia a ficha.
# Ela consulta a Receita Federal (via cnpj.ws) e devolve os campos quebrados
# (cep, logradouro, numero, ie, cnae, etc) num dict pronto pra ser merged
# no `campos` que vai pro SQLite.
#
# FILOSOFIA: a API é um BÔNUS, não um REQUISITO.
# Se ela falhar (timeout, 404, fora do ar), devolve dict com strings vazias
# e o fluxo continua. O robô lida com campos vazios acumulando avisos.
# ══════════════════════════════════════════════════════════════════════════════

def consultar_cnpj_ws(cnpj_limpo):
    """
    Consulta a API cnpj.ws e devolve um dict com os campos enriquecidos.
    Em caso de erro, devolve o mesmo dict mas com todos os valores vazios.
    Nunca levanta exceção — o caller pode confiar que sempre vai receber o dict.
    """
    # Estrutura padrão de retorno (vazia). Garante que o caller sempre tem
    # as chaves esperadas, mesmo se a API falhar.
    dados_vazios = {
        "nome_fantasia":            "",
        "inscricao_estadual":       "",
        "cnae_principal_descricao": "",
        "cep":                      "",
        "logradouro":               "",
        "numero":                   "",
        "complemento":              "",
        "bairro":                   "",
        "uf":                       "",
        "_cnpj_ws_status":          "vazio",  # marcador pra debug
    }

    if not cnpj_limpo or len(cnpj_limpo) != 14:
        print(f"⚠️ [CNPJ.ws] CNPJ inválido ou vazio: {cnpj_limpo!r}")
        return dados_vazios

    try:
        url = f"https://publica.cnpj.ws/cnpj/{cnpj_limpo}"
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=10)

        if resp.status_code != 200:
            print(f"⚠️ [CNPJ.ws] Status {resp.status_code} para CNPJ {cnpj_limpo}")
            dados_vazios["_cnpj_ws_status"] = f"http_{resp.status_code}"
            return dados_vazios

        api_json = resp.json()
        estab = api_json.get("estabelecimento", {}) or {}

        # ── Logradouro: junta tipo + nome (ex: "RUA" + "DAS FLORES") ──
        tipo_log = estab.get("tipo_logradouro", "") or ""
        nome_log = estab.get("logradouro", "") or ""
        logradouro = f"{tipo_log} {nome_log}".strip()

        # ── Complemento: aspirador de espaços (mesma técnica do leitor antigo) ──
        complemento_sujo = estab.get("complemento") or ""
        complemento = " ".join(str(complemento_sujo).split())

        # ── CNAE: junta código + descrição ──
        cnae_id = estab.get("atividade_principal", {}).get("id", "") or ""
        cnae_desc = estab.get("atividade_principal", {}).get("descricao", "") or ""
        cnae_completo = f"{cnae_id} - {cnae_desc}" if cnae_id else cnae_desc

        # ── Inscrição Estadual: pega APENAS as ativas, junta com vírgula ──
        ies_ativas = []
        for ie in estab.get("inscricoes_estaduais", []) or []:
            if ie.get("ativo") is True:
                numero_ie = ie.get("inscricao_estadual")
                if numero_ie:
                    ies_ativas.append(numero_ie)
        ie_final = ", ".join(ies_ativas) if ies_ativas else "Isento / Não encontrada"

        dados_enriquecidos = {
            "nome_fantasia":            estab.get("nome_fantasia", "") or "",
            "inscricao_estadual":       ie_final,
            "cnae_principal_descricao": cnae_completo,
            "cep":                      estab.get("cep", "") or "",
            "logradouro":               logradouro,
            "numero":                   estab.get("numero", "") or "",
            "complemento":              complemento,
            "bairro":                   estab.get("bairro", "") or "",
            "uf":                       (estab.get("estado", {}) or {}).get("sigla", "") or "",
            "_cnpj_ws_status":          "ok",
        }
        print(f"✅ [CNPJ.ws] Dados extraídos para CNPJ {cnpj_limpo}")
        return dados_enriquecidos

    except requests.Timeout:
        print(f"⚠️ [CNPJ.ws] Timeout consultando CNPJ {cnpj_limpo}")
        dados_vazios["_cnpj_ws_status"] = "timeout"
        return dados_vazios

    except requests.ConnectionError as e:
        print(f"⚠️ [CNPJ.ws] Falha de conexão: {e}")
        dados_vazios["_cnpj_ws_status"] = "connection_error"
        return dados_vazios

    except Exception as e:
        print(f"⚠️ [CNPJ.ws] Erro inesperado: {e}")
        dados_vazios["_cnpj_ws_status"] = "erro_inesperado"
        return dados_vazios

app = Flask(__name__, template_folder='templates_reestilizados')
app.secret_key = 'ciamed_super_secreta_2026'

# ← NOVO: Inicializa o banco de cadastros ao subir o servidor
banco_cadastros.init_db()

# Configura pasta de Upload
UPLOAD_FOLDER = 'XML_Entrada'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ==========================================
# PORTEIRO DIGITAL (CONTROLE DE HORÁRIO)
# ==========================================
HORA_INICIO = time(00, 00)
HORA_FIM    = time(23, 59)
DIAS_UTEIS  = [0, 1, 2, 3, 4, 5, 6]

@app.before_request
def verificar_horario():
    if request.endpoint and 'static' in request.endpoint:
        return None

    # ← MODIFICADO: adicionadas as novas rotas liberadas do porteiro
    if request.endpoint in [
        'ficha_cliente', 'api_cnpj_completo', 'consulta_cnpj',
        'confirmar_acao', 'motivo_reprovacao',
        'receber_ficha', 'api_ficha'          # ← NOVAS
    ]:
        return None

    if request.args.get('bypass') == 'leo_admin':
        session['acesso_noturno'] = True
        return redirect(url_for('home'))

    if session.get('acesso_noturno'):
        return None

    agora = datetime.now()
    if agora.weekday() not in DIAS_UTEIS or not (HORA_INICIO <= agora.time() < HORA_FIM):
        return """
        <body style="background-color: #263238; color: white; font-family: 'Segoe UI', sans-serif; display: flex; flex-direction: column; justify-content: center; align-items: center; height: 100vh; margin: 0;">
            <img src="/static/img/ciamedia.png" style="height: 60px; margin-bottom: 20px;">
            <h1 style="color: #05776c; margin-bottom: 10px;">🌙 Sistema em Repouso</h1>
            <p style="font-size: 1.2rem; color: #b0bec5; margin-bottom: 30px; text-align: center; max-width: 500px;">
                A Central de Automação Fiscal da Ciamed opera exclusivamente em horário comercial.
            </p>
            <div style="padding: 20px 40px; border: 1px solid #05776c; border-radius: 8px; background-color: rgba(5, 119, 108, 0.1); text-align: center;">
                <strong style="color: white;">Horário de Funcionamento:</strong><br>
                <span style="color: #b0bec5;">Segunda a Sexta: 08:00 às 18:00</span>
            </div>
            <p style="margin-top: 40px; font-size: 0.85rem; color: #6c757d;">
                Controladoria • Acesso Restrito
            </p>
        </body>
        """, 403

# ==========================================
# ROTAS PRINCIPAIS (FRONTEND & UPLOAD)
# ==========================================

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'status': 'erro', 'msg': 'Nenhum arquivo enviado'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'erro', 'msg': 'Nome vazio'})
    
    if file:
        pasta_upload = app.config['UPLOAD_FOLDER']
        try:
            for arquivo_velho in os.listdir(pasta_upload):
                caminho_velho = os.path.join(pasta_upload, arquivo_velho)
                if os.path.isfile(caminho_velho):
                    os.remove(caminho_velho)
        except Exception as e:
            print(f"⚠️ Erro ao limpar a pasta: {e}")
        
        filepath = os.path.join(pasta_upload, file.filename)
        file.save(filepath)
        return jsonify({'status': 'ok', 'msg': f'Arquivo salvo e fila limpa: {file.filename}'})

# ==========================================
# ROTAS DOS ROBÔS (FASES)
# ==========================================

@app.route('/run_fase1', methods=['GET'])
def rota_fase1():
    meu_id = request.args.get('fila_id', '')
    def gerar_logs():
        tentativas = 0
        while not minha_vez(meu_id):
            tentativas += 1
            if tentativas > 120:
                yield "data: ❌ Timeout na fila. Tente novamente.\n\n"
                sair_da_fila(meu_id)
                return
            yield f"data: [FILA] posição {posicao_na_fila(meu_id)}\n\n"
            time_module.sleep(1)
        iniciar_execucao(meu_id)
        yield "data: [FILA_LIBERADA]\n\n"
        try:
            comando = [sys.executable, "-X", "utf8", "-u", "automacoes/fase1_ncm.py"]
            processo = subprocess.Popen(comando, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', bufsize=1)
            for linha in iter(processo.stdout.readline, ''):
                linha_limpa = linha.strip()
                if not linha_limpa: linha_limpa = " "
                yield f"data: {linha_limpa}\n\n"
            processo.stdout.close()
            processo.wait()
            yield "data: [FIM_DO_PROCESSO]\n\n"
        finally:
            finalizar_execucao()
    return Response(gerar_logs(), mimetype='text/event-stream')

@app.route('/run_fase2', methods=['GET'])
def run_fase2():
    meu_id = request.args.get('fila_id', '')
    def gerar_logs():
        tentativas = 0
        while not minha_vez(meu_id):
            tentativas += 1
            if tentativas > 120:
                yield "data: ❌ Timeout na fila. Tente novamente.\n\n"
                sair_da_fila(meu_id)
                return
            yield f"data: [FILA] posição {posicao_na_fila(meu_id)}\n\n"
            time_module.sleep(1)
        iniciar_execucao(meu_id)
        yield "data: [FILA_LIBERADA]\n\n"
        try:
            comando = [sys.executable, "-X", "utf8", "-u", "automacoes/fase2_item.py"]
            processo = subprocess.Popen(comando, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', bufsize=1)
            for linha in iter(processo.stdout.readline, ''):
                linha_limpa = linha.strip()
                if not linha_limpa: linha_limpa = " "
                yield f"data: {linha_limpa}\n\n"
            processo.stdout.close()
            processo.wait()
            yield "data: [FIM_DO_PROCESSO]\n\n"
        finally:
            finalizar_execucao()
    return Response(gerar_logs(), mimetype='text/event-stream')

@app.route('/run_mdf', methods=['GET'])
def rota_mdf():
    meu_id     = request.args.get('fila_id', '')
    data_inicio = request.args.get('inicio', '01/01/2026')
    data_fim    = request.args.get('fim', '31/01/2026')
    def gerar_logs():
        tentativas = 0
        while not minha_vez(meu_id):
            tentativas += 1
            if tentativas > 120:
                yield "data: ❌ Timeout na fila. Tente novamente.\n\n"
                sair_da_fila(meu_id)
                return
            yield f"data: [FILA] posição {posicao_na_fila(meu_id)}\n\n"
            time_module.sleep(1)
        iniciar_execucao(meu_id)
        yield "data: [FILA_LIBERADA]\n\n"
        try:
            processo = subprocess.Popen(['python', '-u', 'automacoes/robo_mdf.py', data_inicio, data_fim], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, encoding='utf-8', bufsize=1)
            for linha in iter(processo.stdout.readline, ''):
                if linha: yield f"data: {linha.strip()}\n\n"
            processo.stdout.close()
            processo.wait()
            yield "data: [FIM_DO_PROCESSO]\n\n"
        finally:
            finalizar_execucao()
    return Response(gerar_logs(), mimetype='text/event-stream')

@app.route('/download_mdf')
def download_mdf():
    caminho_arquivo = os.path.join(os.getcwd(), "Relatorio_MDFs_Gerado.xlsx")
    if not os.path.exists(caminho_arquivo):
        return "❌ O arquivo não foi encontrado. O robô pode não ter encontrado dados neste período ou ocorreu um erro na geração.", 404
    try:
        return send_file(caminho_arquivo, as_attachment=True, download_name="Relatorio_MDF_Ciamed.xlsx")
    except Exception as e:
        return f"Erro interno ao tentar baixar o arquivo: {e}", 500

@app.route('/cadastro_fornecedor')
def cadastro_fornecedor():
    return render_template('cadastro_fornecedor.html')

@app.route('/cadastro_itens')
def cadastro_itens():
    return render_template('cadastro_itens.html')

@app.route('/relatorio_mdf')
def relatorio_mdf():
    return render_template('relatorio_mdf.html')

@app.route('/run_fornecedor')
def run_fornecedor():
    meu_id    = request.args.get('fila_id', '')
    dados_json = request.args.get('dados', '{}')
    def generate():
        tentativas = 0
        while not minha_vez(meu_id):
            tentativas += 1
            if tentativas > 120:
                yield "data: ❌ Timeout na fila. Tente novamente.\n\n"
                sair_da_fila(meu_id)
                return
            yield f"data: [FILA] posição {posicao_na_fila(meu_id)}\n\n"
            time_module.sleep(1)
        iniciar_execucao(meu_id)
        yield "data: [FILA_LIBERADA]\n\n"
        try:
            process = subprocess.Popen([sys.executable, 'automacoes/robo_fornecedor.py', dados_json], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1, encoding='utf-8')
            for line in process.stdout:
                linha = line.strip()
                if linha: yield f"data: {linha}\n\n"
            process.wait()
            yield "data: [FIM_DO_PROCESSO]\n\n"
        except Exception as e:
            yield f"data: ❌ Erro interno: {str(e)}\n\n"
            yield "data: [FIM_DO_PROCESSO]\n\n"
        finally:
            finalizar_execucao()
    return Response(generate(), mimetype='text/event-stream')

@app.route('/cadastro_clientes')
def cadastro_clientes():
    return render_template('cadastro_clientes.html')


@app.route('/ficha_cliente')
def ficha_cliente():
    return render_template('ficha_cliente.html')


# ══════════════════════════════════════════════════════════════════════════════
# ROTA NOVA: Receber ficha do cliente
# ──────────────────────────────────────────────────────────────────────────────
# O JS agora envia para cá em vez de direto pro N8N.
# O Flask:
#   1. Salva os arquivos em uploads_fichas/{uuid}/
#   2. Salva os dados no SQLite
#   3. Compara documentos (quando for correção)
#   4. Repassa para o N8N via JSON com todos os metadados
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/receber_ficha', methods=['POST'])
def receber_ficha():

    # ── 1. Verificar se é correção (tem uuid no form) ──────────────────────
    uuid_existente = request.form.get('uuid', '').strip()
    is_correcao    = bool(uuid_existente)

    if is_correcao:
        sub_original = banco_cadastros.buscar_submissao(uuid_existente)
        if not sub_original:
            return jsonify({'status': 'erro', 'mensagem': 'UUID inválido ou expirado.'}), 400
        submission_uuid = uuid_existente
        docs_originais  = sub_original['docs_enviados']
    else:
        submission_uuid = str(uuid_lib.uuid4())
        docs_originais  = []

    # ── 2. Extrair campos de texto ─────────────────────────────────────────
    campos = {
        'vendedor':               request.form.get('vendedor',               ''),
        'representante':          request.form.get('representante',           ''),
        'captacao':               request.form.get('captacao',                ''),
        'tipo_cadastro':          request.form.get('tipo_cadastro',           ''),
        'codigo_nl':              request.form.get('codigo_nl',               ''),
        'contribuinte_icms':      request.form.get('contribuinte_icms',       ''),
        'possui_regime_especial': request.form.get('possui_regime_especial',  ''),
        'regra_faturamento':      request.form.get('regra_faturamento',       ''),
        'cnpj':                   request.form.get('cnpj',                    ''),
        'razao_social':           request.form.get('razao_social',            ''),
        'endereco':               request.form.get('endereco',                ''),
        'telefone_empresa':       request.form.get('telefone_empresa',        ''),
        'whatsapp':               request.form.get('whatsapp',                ''),
        'email_xml_1':            request.form.get('email_xml_1',             ''),
        'email_xml_2':            request.form.get('email_xml_2',             ''),
        'email_xml_3':            request.form.get('email_xml_3',             ''),
        'compras_nome':           request.form.get('compras_nome',            ''),
        'compras_tel':            request.form.get('compras_tel',             ''),
        'compras_email':          request.form.get('compras_email',           ''),
        'rec_nome':               request.form.get('rec_nome',                ''),
        'rec_tel':                request.form.get('rec_tel',                 ''),
        'rec_email':              request.form.get('rec_email',               ''),
        'fin_nome':               request.form.get('fin_nome',                ''),
        'fin_tel':                request.form.get('fin_tel',                 ''),
        'fin_email':              request.form.get('fin_email',               ''),
        'farma_nome':             request.form.get('farma_nome',              ''),
        'farma_tel':              request.form.get('farma_tel',               ''),
        'farma_email':            request.form.get('farma_email',             ''),
    }

    # ── 2.1. ENRIQUECIMENTO via CNPJ.ws ────────────────────────────────────
    # Consulta a Receita Federal pra preencher campos que a ficha digital
    # não captura (cep, logradouro separado, IE, CNAE, etc). Roda "por
    # debaixo dos panos" — o cliente nem sabe que isso aconteceu.
    #
    # IMPORTANTE: essa consulta NUNCA bloqueia o fluxo. Se a API falhar,
    # os campos enriquecidos ficam vazios e o robô acumula avisos depois.
    cnpj_limpo = re.sub(r'\D', '', campos['cnpj'])
    dados_cnpj_ws = consultar_cnpj_ws(cnpj_limpo)

    # Faz o merge: tudo que veio da CNPJ.ws entra no `campos` com prefixo
    # claro pra não conflitar com nada que já existe ali. Quando a rota
    # /n8n_iniciar_reativacao for criada, ela vai ler esses campos.
    campos['cnpj_limpo']               = cnpj_limpo
    campos['nome_fantasia']            = dados_cnpj_ws['nome_fantasia']
    campos['inscricao_estadual']       = dados_cnpj_ws['inscricao_estadual']
    campos['cnae_principal_descricao'] = dados_cnpj_ws['cnae_principal_descricao']
    campos['cep']                      = dados_cnpj_ws['cep']
    campos['logradouro']               = dados_cnpj_ws['logradouro']
    campos['numero']                   = dados_cnpj_ws['numero']
    campos['complemento']              = dados_cnpj_ws['complemento']
    campos['bairro']                   = dados_cnpj_ws['bairro']
    campos['uf']                       = dados_cnpj_ws['uf']
    campos['_cnpj_ws_status']          = dados_cnpj_ws['_cnpj_ws_status']

    
    # ── 3. Salvar arquivos em uploads_fichas/{uuid}/ ───────────────────────
    lista_docs  = [
        'doc_regime_especial', 'doc_alvara', 'doc_crt', 'doc_afe',
        'doc_balanco', 'doc_balancete', 'doc_contrato'
    ]
    pasta_uuid  = os.path.join('uploads_fichas', submission_uuid)
    os.makedirs(pasta_uuid, exist_ok=True)

    docs_novos   = []  # docs com NOVO arquivo nesta submissão
    docs_manter  = json.loads(request.form.get('docs_manter', '[]'))

    for doc_nome in lista_docs:
        arquivo = request.files.get(doc_nome)
        if arquivo and arquivo.filename:
            nome_seguro = secure_filename(arquivo.filename)
            ext         = os.path.splitext(nome_seguro)[1].lower() or '.pdf'
            caminho     = os.path.join(pasta_uuid, f"{doc_nome}{ext}")
            arquivo.save(caminho)
            docs_novos.append(doc_nome)

    # ── 4. Calcular docs finais disponíveis ───────────────────────────────
    # docs_manter são os que o vendedor NÃO resubmetiu mas existiam antes
    docs_mantidos_validos = [d for d in docs_manter if d in docs_originais]
    docs_finais           = list(set(docs_novos + docs_mantidos_validos))

    # ── 5. Categorizar documentos (só relevante na correção) ───────────────
    if is_correcao:
        # 📎 Já existentes: estavam no original E não foram resubmetidos
        docs_ja_existentes = [d for d in docs_mantidos_validos]
        # 🔄 Substituídos: estavam no original E foram resubmetidos com arquivo novo
        docs_substituidos  = [d for d in docs_novos if d in docs_originais]
        # 🆕 Adicionados: NÃO estavam no original E foram enviados agora
        docs_adicionados   = [d for d in docs_novos if d not in docs_originais]

        nova_tentativa = banco_cadastros.incrementar_tentativa(
            submission_uuid, campos, docs_finais
        )
    else:
        docs_ja_existentes = []
        docs_substituidos  = []
        docs_adicionados   = []
        nova_tentativa     = 1

        banco_cadastros.salvar_submissao(
            submission_uuid,
            campos['cnpj'],
            campos['razao_social'],
            campos,
            docs_finais
        )

    # ── 6. Montar nomes legíveis dos docs para exibir nos emails ──────────
    # O N8N vai usar estas strings diretamente nos emails da farmacêutica
    nomes_legíveis = {
        'doc_regime_especial': 'Regime Especial de ICMS',
        'doc_alvara':          'Alvará Sanitário',
        'doc_crt':             'Certidão de Regularidade',
        'doc_afe':             'Publicação AFE / AE',
        'doc_balanco':         'Último Balanço e DRE',
        'doc_balancete':       'Balancete (< 90 dias)',
        'doc_contrato':        'Contrato Social',
    }

    def nomes(lista):
        return ', '.join(nomes_legíveis.get(d, d) for d in lista) if lista else '—'

    # ── 7. Montar payload para o N8N ──────────────────────────────────────
    payload_n8n = {
        **campos,
        'uuid':               submission_uuid,
        'tentativa':          nova_tentativa,
        'is_correcao':        str(is_correcao).lower(),
        'docs_ja_existentes': nomes(docs_ja_existentes),
        'docs_substituidos':  nomes(docs_substituidos),
        'docs_adicionados':   nomes(docs_adicionados),
    }

    # ── 8. Disparar para o Webhook do N8N (multipart com arquivos) ────────
    # Substitua a string inteira por isto:
    N8N_WEBHOOK_URL = os.getenv('N8N_WEBHOOK_CADASTRO')

    payload_str = {k: str(v) for k, v in payload_n8n.items()}

    arquivos_abertos = {}
    pasta = os.path.join('uploads_fichas', submission_uuid)

    if os.path.exists(pasta):
        for arquivo in os.listdir(pasta):
            nome_sem_ext = os.path.splitext(arquivo)[0]
            extensao     = os.path.splitext(arquivo)[1]
            caminho      = os.path.join(pasta, arquivo)
            if nome_sem_ext in docs_finais:
                arquivos_abertos[nome_sem_ext] = (
                    arquivo,
                    open(caminho, 'rb'),
                    'application/octet-stream'
                )

    try:
        resposta_n8n = requests.post(
            N8N_WEBHOOK_URL,
            data=payload_str,
            files=arquivos_abertos if arquivos_abertos else None,
            timeout=60
        )
        print(f"✅ [N8N] Status: {resposta_n8n.status_code} | UUID: {submission_uuid} | Tentativa: {nova_tentativa} | Docs: {list(arquivos_abertos.keys())}")
        return jsonify({'status': 'ok', 'uuid': submission_uuid})

    except Exception as e:
        print(f"❌ [N8N] Erro ao disparar webhook: {e}")
        return jsonify({'status': 'erro', 'mensagem': str(e)}), 500

    finally:
        for nome, (_, fp, _) in arquivos_abertos.items():
            try:
                fp.close()
            except Exception:
                pass

# ══════════════════════════════════════════════════════════════════════════════
# ROTA NOVA: Retornar dados de uma submissão para pré-preenchimento
# ──────────────────────────────────────────────────────────────────────────────
# O JS chama esta rota quando detecta ?correcao=UUID na URL da ficha
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/ficha/<submission_uuid>')
def api_ficha(submission_uuid):
    sub = banco_cadastros.buscar_submissao(submission_uuid)
    if not sub:
        return jsonify({'erro': 'Submissão não encontrada'}), 404

    return jsonify({
        'dados':              sub['dados_json'],
        'docs_enviados':      sub['docs_enviados'],
        'motivos_reprovacao': sub['motivos_reprovacao'],
        'tentativa':          sub['tentativa'],
        'razao_social':       sub['razao_social'],
        'cnpj':               sub['cnpj'],
    })


# ──────────────────────────────────────────────────────────────────────────────
# ROTA: Farmacêutica clica no botão do email → Flask chama N8N e mostra tela bonita
#
# Uso nos emails do N8N:
#   APROVAR:  /confirmar?acao=aprovar&empresa=NOME&resumeUrl=URL_N8N
#   MÁSCARAS: /confirmar?acao=mascaras&empresa=NOME&resumeUrl=URL_N8N
#   REPROVAR: → vai para /motivo_reprovacao (não passa por aqui diretamente)
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/confirmar')
def confirmar_acao():
    acao       = request.args.get('acao',      '')
    empresa    = request.args.get('empresa',   '')
    resume_url = request.args.get('resumeUrl', '')
    motivo     = request.args.get('motivo',    '')
    uuid_ficha = request.args.get('uuid',      '')

    if not resume_url:
        return "Parâmetro resumeUrl ausente.", 400

    try:
        if acao == 'aprovar':
            req_ext.get(resume_url, params={'decisao': 'aprovar'}, timeout=10)
            if uuid_ficha:
                banco_cadastros.atualizar_status(uuid_ficha, 'aprovado')

        elif acao == 'reprovar':
            req_ext.get(resume_url, params={'decisao': 'reprovar', 'motivo': motivo}, timeout=10)
            if uuid_ficha and motivo:
                banco_cadastros.registrar_reprovacao(uuid_ficha, motivo)

        elif acao == 'mascaras':
            req_ext.get(resume_url, params={'mascaras': 'confirmadas'}, timeout=10)

    except Exception as e:
        print(f"⚠️ Aviso ao chamar N8N: {e}")

    return render_template('confirmar_acao.html', acao=acao, empresa=empresa, motivo=motivo)


# ──────────────────────────────────────────────────────────────────────────────
# ROTA: Formulário de motivo de reprovação
# Farmacêutica clica "Reprovar" → preenche motivo → vai para /confirmar
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/motivo_reprovacao')
def motivo_reprovacao():
    resume_url = request.args.get('resumeUrl', '')
    empresa    = request.args.get('empresa',   '')
    cnpj       = request.args.get('cnpj',      '')
    uuid_ficha = request.args.get('uuid',      '')   # ← NOVO
    return render_template(
        'motivo_reprovacao.html',
        resume_url=resume_url,
        empresa=empresa,
        cnpj=cnpj,
        uuid=uuid_ficha                               # ← NOVO
    )


# ==========================================
# SISTEMA DE BANCO DE DADOS (JSON) — Chamados
# ==========================================
ARQUIVO_CHAMADOS = 'dados/chamados.json'

def carregar_chamados():
    if not os.path.exists(ARQUIVO_CHAMADOS):
        return []
    with open(ARQUIVO_CHAMADOS, 'r', encoding='utf-8') as f:
        return json.load(f)

def salvar_chamados(chamados):
    with open(ARQUIVO_CHAMADOS, 'w', encoding='utf-8') as f:
        json.dump(chamados, f, indent=4, ensure_ascii=False)

# ==========================================
# ROTAS DO ADMIN E DO SUPORTE
# ==========================================

@app.route('/suporte')
def suporte():
    return render_template('suporte.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/enviar_chamado', methods=['POST'])
def enviar_chamado():
    try:
        dados = request.get_json()
        novo_chamado = {
            "id":       int(datetime.now().timestamp()),
            "nome":     dados.get('nome'),
            "setor":    dados.get('setor'),
            "assunto":  dados.get('assunto'),
            "mensagem": dados.get('mensagem'),
            "data":     datetime.now().strftime("%d/%m/%Y %H:%M"),
            "status":   "pendente"
        }
        chamados = carregar_chamados()
        chamados.append(novo_chamado)
        salvar_chamados(chamados)
        return jsonify({'status': 'ok'})
    except Exception as e:
        print(f"ERRO AO SALVAR CHAMADO: {str(e)}")
        return jsonify({'status': 'erro', 'msg': 'Erro interno ao salvar no banco.'})

@app.route('/api/chamados', methods=['GET'])
def get_chamados():
    return jsonify(carregar_chamados())

@app.route('/api/chamados/<int:id>/status', methods=['POST'])
def update_status(id):
    novo_status = request.get_json().get('status')
    chamados    = carregar_chamados()
    for c in chamados:
        if c['id'] == id:
            c['status'] = novo_status
            break
    salvar_chamados(chamados)
    return jsonify({'status': 'ok'})

@app.route('/api/chamados/<int:id>', methods=['DELETE'])
def delete_chamado(id):
    chamados = carregar_chamados()
    chamados = [c for c in chamados if c['id'] != id]
    salvar_chamados(chamados)
    return jsonify({'status': 'ok'})

@app.route('/consulta_cnpj/<cnpj>')
def consulta_cnpj(cnpj):
    try:
        headers  = {'User-Agent': 'Mozilla/5.0'}
        url      = f'https://publica.cnpj.ws/cnpj/{cnpj}'
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 200:
            data   = response.json()
            estabs = data.get('estabelecimento', {})
            uf_fornecedor = estabs.get('estado', {}).get('sigla')
            lista_ie      = estabs.get('inscricoes_estaduais', [])
            ie_ativa      = ""
            for item in lista_ie:
                if item.get('estado', {}).get('sigla') == uf_fornecedor and item.get('ativo') is True:
                    ie_ativa = item.get('inscricao_estadual')
                    break
            simples    = data.get('simples') or {}
            status_mei = "SIM" if simples.get('mei') == 'Sim' else "NÃO"
            resultado  = {
                'razao_social':    data.get('razao_social'),
                'nome_fantasia':   estabs.get('nome_fantasia'),
                'cnpj':            cnpj,
                'inscricao_estadual': ie_ativa,
                'mei':             status_mei,
                'logradouro':      f"{estabs.get('tipo_logradouro', '')} {estabs.get('logradouro', '')}".strip(),
                'numero':          estabs.get('numero'),
                'complemento':     estabs.get('complemento'),
                'bairro':          estabs.get('bairro'),
                'cep':             estabs.get('cep'),
                'uf':              uf_fornecedor,
                'municipio':       estabs.get('cidade', {}).get('nome'),
                'cnae_fiscal':     estabs.get('atividade_principal', {}).get('id'),
                'cnae_fiscal_descricao': estabs.get('atividade_principal', {}).get('descricao'),
                'data_inicio_atividade': estabs.get('data_inicio_atividade'),
                'descricao_identificador_matriz_filial': "MATRIZ" if estabs.get('tipo') == "Matriz" else "FILIAL",
                'codigo_natureza_juridica': data.get('natureza_juridica', {}).get('id'),
                'natureza_juridica': data.get('natureza_juridica', {}).get('descricao')
            }
            return jsonify(resultado)
        else:
            msg = "CNPJ não encontrado ou erro na API."
            if response.status_code == 429: msg = "Muitas consultas! Aguarde 1 minuto."
            return jsonify({'erro': True, 'message': msg}), response.status_code
    except Exception as e:
        return jsonify({'erro': True, 'message': str(e)}), 500

@app.route('/api/cnpj_completo/<cnpj>')
def api_cnpj_completo(cnpj):
    try:
        cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')
        headers    = {'User-Agent': 'Mozilla/5.0'}
        url        = f'https://publica.cnpj.ws/cnpj/{cnpj_limpo}'
        response   = requests.get(url, headers=headers, timeout=45)
        if response.status_code == 200:
            return jsonify(response.json())
        else:
            msg = "CNPJ não encontrado."
            if response.status_code == 429:
                msg = "Muitas consultas! Aguarde 1 minuto e tente novamente."
            return jsonify({'erro': True, 'message': msg}), response.status_code
    except requests.exceptions.Timeout:
        return jsonify({'erro': True, 'message': "A API da Receita Federal demorou muito. Tente novamente!"}), 504
    except Exception as e:
        return jsonify({'erro': True, 'message': f"Erro de comunicação: {str(e)}"}), 500

@app.route('/consulta_cnpj_completa')
def consulta_cnpj_completa():
    return render_template('consulta_cnpj.html')

@app.route('/fila/entrar', methods=['POST'])
def fila_entrar():
    meu_id = entrar_na_fila()
    return jsonify({'id': meu_id, 'posicao': posicao_na_fila(meu_id)})

@app.route('/fila/status/<meu_id>')
def fila_status(meu_id):
    pos = posicao_na_fila(meu_id)
    return jsonify({
        'posicao':     pos,
        'minha_vez':   pos == 1 and _em_execucao is None,
        'em_execucao': _em_execucao is not None
    })

@app.route('/fila/sair/<meu_id>', methods=['POST'])
def fila_sair(meu_id):
    sair_da_fila(meu_id)
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════
# FILA DO ROBÔ DE CADASTRO — controle via N8N
# ══════════════════════════════════════════════════════════════════

# Fila específica para o robô de cadastro
# Cada item é um dict: { 'id': str, 'resume_url': str }
_fila_cadastro = []
_fila_cadastro_lock = threading.Lock()
_cadastro_em_execucao = False

@app.route('/fila_cadastro/entrar', methods=['POST'])
def fila_cadastro_entrar():
    """
    N8N chama aqui para entrar na fila do robô.
    Recebe: { "resume_url": "https://n8n.cloud/webhook/..." }
    Retorna: { "id": "...", "posicao": N, "sua_vez": true/false }
    """
    global _cadastro_em_execucao

    dados      = request.get_json()
    resume_url = dados.get('resume_url', '')

    if not resume_url:
        return jsonify({'erro': 'resume_url obrigatório'}), 400

    # Gera um ID único para esta execução
    novo_id = str(uuid_lib.uuid4())[:8]

    with _fila_cadastro_lock:
        _fila_cadastro.append({
            'id':         novo_id,
            'resume_url': resume_url
        })
        posicao    = len(_fila_cadastro)
        sua_vez    = posicao == 1 and not _cadastro_em_execucao

    print(f"📥 [FILA] Entrada | ID: {novo_id} | Posição: {posicao}")

    # Se for o primeiro e ninguém está executando, libera imediatamente
    if sua_vez:
        _liberar_proximo_cadastro()

    return jsonify({
        'id':      novo_id,
        'posicao': posicao,
        'sua_vez': sua_vez
    })


@app.route('/fila_cadastro/liberar_proximo', methods=['POST'])
def fila_cadastro_liberar_proximo():
    """
    N8N chama aqui quando o robô terminou.
    Remove o atual da fila e acorda o próximo Wait node.
    """
    global _cadastro_em_execucao

    with _fila_cadastro_lock:
        # Remove o primeiro (que acabou de terminar)
        if _fila_cadastro:
            concluido = _fila_cadastro.pop(0)
            print(f"✅ [FILA] Concluído | ID: {concluido['id']}")
        _cadastro_em_execucao = False

    # Acorda o próximo se existir
    _liberar_proximo_cadastro()

    return jsonify({'status': 'ok'})


def _liberar_proximo_cadastro():
    """
    Função interna: pega o primeiro da fila e chama o resumeUrl dele
    no N8N para acordar o Wait node correspondente.
    """
    global _cadastro_em_execucao

    with _fila_cadastro_lock:
        if not _fila_cadastro or _cadastro_em_execucao:
            return
        proximo        = _fila_cadastro[0]
        _cadastro_em_execucao = True

    print(f"🚀 [FILA] Liberando | ID: {proximo['id']}")

    # Chama o resumeUrl do Wait node do N8N em background
    # (em thread separada para não travar o Flask)
    def chamar_n8n():
        try:
            req_ext.get(proximo['resume_url'], timeout=10)
            print(f"✅ [FILA] N8N acordado | ID: {proximo['id']}")
        except Exception as e:
            print(f"⚠️ [FILA] Erro ao acordar N8N: {e}")

    threading.Thread(target=chamar_n8n, daemon=True).start()

# ==========================================
# ROTA DE INTEGRAÇÃO: n8n -> ROBÔ RPA
# ==========================================
@app.route('/n8n_iniciar_cadastro', methods=['POST'])
def n8n_iniciar_cadastro():
    dados_n8n = request.json
    cnpj      = dados_n8n.get('cnpj')
    razao     = dados_n8n.get('razao_social')
    vendedor  = dados_n8n.get('vendedor')
    print(f"\n🔔 [N8N] Solicitação recebida!")
    print(f"🏢 Cliente: {razao} | 👤 Vendedor: {vendedor}")
    try:
        resultado = cadastro_novo.executar({
            "cnpj": cnpj,
            "razao_social": razao,
            "vendedor": vendedor
        })
        return jsonify({
            "status":    "sucesso",
            "mensagem":  f"Robô finalizou o cadastro de {razao}",
            "resultado": resultado
        }), 200
    except Exception as e:
        print(f"❌ Erro ao processar chamado do n8n: {e}")
        return jsonify({"status": "erro", "detalhes": str(e)}), 500
    
# ══════════════════════════════════════════════════════════════════════════════
# ADICIONAR ESTAS ROTAS NO app.py ANTES DA LINHA:
# if __name__ == '__main__':
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/monitor_cadastros')
def monitor_cadastros():
    return render_template('monitor_cadastros.html')


@app.route('/api/monitor_cadastros')
def api_monitor_cadastros():
    """
    API que alimenta a tela de Monitor de Cadastros.
    Aceita query params: status, tipo, busca, data_de, data_ate
    Retorna: { cadastros: [...], stats: {...} }
    """
    status   = request.args.get('status',   None)
    tipo     = request.args.get('tipo',     None)
    busca    = request.args.get('busca',    None)
    data_de  = request.args.get('data_de',  None)
    data_ate = request.args.get('data_ate', None)

    cadastros = banco_cadastros.listar_submissoes(
        status=status,
        tipo=tipo,
        busca=busca,
        data_de=data_de,
        data_ate=data_ate
    )
    stats = banco_cadastros.stats_submissoes()

    return jsonify({
        'cadastros': cadastros,
        'stats':     stats
    })

# ──────────────────────────────────────────────────────────────────────────────
# ROTA: Baixar/Visualizar documentos direto do Monitor de Cadastros
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/download_doc/<uuid>/<doc_nome>')
def download_doc(uuid, doc_nome):
    # Acessa a pasta específica deste cadastro usando o UUID
    pasta_cliente = os.path.join(os.getcwd(), 'uploads_fichas', uuid)
    
    # Verifica se a pasta existe
    if os.path.exists(pasta_cliente):
        # Varre os arquivos da pasta para encontrar a extensão correta (.pdf, .jpg, etc)
        for arquivo in os.listdir(pasta_cliente):
            nome_sem_extensao = os.path.splitext(arquivo)[0]
            
            if nome_sem_extensao == doc_nome:
                caminho_completo = os.path.join(pasta_cliente, arquivo)
                # as_attachment=False faz o navegador TENTAR ABRIR na tela (ex: PDF) ao invés de só baixar
                return send_file(caminho_completo, as_attachment=False)
                
    return "❌ Documento não encontrado no servidor. Ele pode ter sido apagado ou não foi salvo corretamente.", 404


# ══════════════════════════════════════════════════════════════════════════════
# ROTA: Exportar fichas cadastrais como .xlsx formatado
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/exportar_fichas')
def exportar_fichas():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Buscar dados com os mesmos filtros do Monitor ─────────────────────────
    registros = banco_cadastros.listar_submissoes(
        status   = request.args.get('status',   None),
        tipo     = request.args.get('tipo',     None),
        busca    = request.args.get('busca',    None),
        data_de  = request.args.get('data_de',  None),
        data_ate = request.args.get('data_ate', None)
    )

    # ── Configurar workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Fichas Cadastrais"

    # Cores
    COR_TITULO_BG  = "1A2744"
    COR_TITULO_FT  = "FFFFFF"
    COR_HEADER_BG  = "0EE6B8"
    COR_HEADER_FT  = "090D14"
    COR_LINHA_PAR  = "EEF4FF"
    COR_LINHA_IMP  = "FFFFFF"
    COR_STATUS = {
        "pendente":   ("FEF3C7","92400E"),
        "aprovado":   ("D1FAE5","065F46"),
        "reprovado":  ("FEE2E2","991B1B"),
        "cadastrado": ("EDE9FE","5B21B6"),
    }

    def fill(cor):
        return PatternFill("solid", start_color=cor, end_color=cor)

    def borda():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def cel_fmt(cell, bold=False, cor_ft="000000", cor_bg=None, size=9, wrap=False, h_align="left"):
        cell.font      = Font(name="Arial", bold=bold, color=cor_ft, size=size)
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        cell.border    = borda()
        if cor_bg:
            cell.fill = fill(cor_bg)

    # ── Linha 1: título do relatório ──────────────────────────────────────────
    ws.merge_cells("A1:R1")
    ws["A1"] = "CIAMED  —  Relatório de Fichas Cadastrais"
    ws["A1"].font      = Font(name="Arial", bold=True, size=15, color=COR_TITULO_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill      = fill("F0F4FF")
    ws.row_dimensions[1].height = 30

    # ── Linha 2: subtítulo com data e filtros ─────────────────────────────────
    from datetime import datetime as _dt
    params = request.args
    partes = []
    if params.get("status"):   partes.append(f"Status: {params['status']}")
    if params.get("tipo"):     partes.append(f"Tipo: {params['tipo']}")
    if params.get("busca"):    partes.append(f"Busca: {params['busca']}")
    if params.get("data_de"):  partes.append(f"De: {params['data_de']}")
    if params.get("data_ate"): partes.append(f"Até: {params['data_ate']}")
    filtros_str = "  |  ".join(partes) if partes else "Sem filtros — todos os registros"

    ws.merge_cells("A2:R2")
    ws["A2"] = f"Gerado em: {_dt.now().strftime('%d/%m/%Y às %H:%M')}     •     Filtros: {filtros_str}     •     Total: {len(registros)} ficha(s)"
    ws["A2"].font      = Font(name="Arial", size=9, color="555555", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill      = fill("F0F4FF")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # espaçamento

    # ── Linha 4: cabeçalhos das colunas ──────────────────────────────────────
    COLUNAS = [
        ("Data/Hora",             18),
        ("Razão Social",          32),
        ("CNPJ",                  17),
        ("Tipo",                  14),
        ("Status",                14),
        ("Tentativa",              9),
        ("Cód. ERP",              12),
        ("Vendedor",              15),
        ("Representante",         16),
        ("Captação",              15),
        ("ICMS Contribuinte",     15),
        ("Regime Especial",       15),
        ("Regra Faturamento",     15),
        ("Telefone",              14),
        ("WhatsApp",              14),
        ("E-mail XML",            28),
        ("Documentos",            38),
        ("Motivo Reprovação",     42),
    ]

    NOMES_DOCS = {
        "doc_regime_especial": "Regime Especial ICMS",
        "doc_alvara":          "Alvará Sanitário",
        "doc_crt":             "Certidão de Regularidade",
        "doc_afe":             "AFE / AE",
        "doc_balanco":         "Balanço e DRE",
        "doc_balancete":       "Balancete",
        "doc_contrato":        "Contrato Social",
    }

    LIN_HEADER = 4
    for col_idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        letra = get_column_letter(col_idx)
        cell  = ws.cell(row=LIN_HEADER, column=col_idx, value=titulo)
        cel_fmt(cell, bold=True, cor_ft=COR_HEADER_FT, cor_bg=COR_HEADER_BG,
                size=9, h_align="center")
        ws.column_dimensions[letra].width = largura
    ws.row_dimensions[LIN_HEADER].height = 22

    # ── Linhas de dados ───────────────────────────────────────────────────────
    LABEL_STATUS = {
        "pendente":   "⏳ Em Análise",
        "aprovado":   "✅ Aprovado",
        "reprovado":  "❌ Reprovado",
        "cadastrado": "🤖 Cadastrado ERP",
    }

    for i, c in enumerate(registros):
        lin   = LIN_HEADER + 1 + i
        d     = c.get("dados_json") or {}
        status= c.get("status", "pendente")
        bg_lin= COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMP

        # Documentos como lista legível
        docs_lista = [NOMES_DOCS.get(x, x) for x in (c.get("docs_enviados") or [])]
        docs_str   = ", ".join(docs_lista) if docs_lista else "—"

        # Motivo mais recente
        motivos    = c.get("motivos_reprovacao") or []
        motivo_str = motivos[-1]["motivo"] if motivos else "—"

        valores = [
            c.get("created_at", "—"),
            c.get("razao_social", "—"),
            c.get("cnpj", "—"),
            d.get("tipo_cadastro", "—"),
            LABEL_STATUS.get(status, status),
            f"{c.get('tentativa', 1)}ª",
            d.get("cod_erp", "—"),
            d.get("vendedor", "—"),
            d.get("representante", "—"),
            d.get("captacao", "—"),
            d.get("contribuinte_icms", "—"),
            d.get("possui_regime_especial", "—"),
            d.get("regra_faturamento", "—"),
            d.get("telefone_empresa", "—"),
            d.get("whatsapp", "—"),
            d.get("email_xml_1", "—"),
            docs_str,
            motivo_str,
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=lin, column=col_idx, value=valor)
            h_al = "center" if col_idx in [4,5,6,7] else "left"

            # Cor especial na coluna Status
            if col_idx == 5:
                bg_s, ft_s = COR_STATUS.get(status, (bg_lin, "000000"))
                cel_fmt(cell, cor_ft=ft_s, cor_bg=bg_s, size=9, h_align="center", bold=True)
            else:
                cel_fmt(cell, cor_ft="222222", cor_bg=bg_lin, size=9, h_align=h_al, wrap=(col_idx in [16,17,18]))

        ws.row_dimensions[lin].height = 18

    # ── Congelar painel no cabeçalho ──────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Auto-filtro no cabeçalho ──────────────────────────────────────────────
    ultima_col = get_column_letter(len(COLUNAS))
    ws.auto_filter.ref = f"A{LIN_HEADER}:{ultima_col}{LIN_HEADER + len(registros)}"

    # ── Exportar para memória e devolver como download ────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"Fichas_Cadastrais_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    print("🚀 SERVIDOR ONLINE! Acesse pelo IP da sua máquina.")
    app.run(host='0.0.0.0', port=5000, debug=True)