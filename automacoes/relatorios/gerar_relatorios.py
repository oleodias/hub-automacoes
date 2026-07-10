# ══════════════════════════════════════════════════════════════
# gerar_relatorios.py — Robô RPA dos relatórios aos laboratórios
# ──────────────────────────────────────────────────────────────
# Gera no NL (NLWeb/APEX) os dois relatórios — Análise de Estoque e
# Demanda Fornecedor — baixa em CSV, converte para .xlsx e devolve um
# MANIFESTO mapeando lab_id -> arquivos. O envio dos e-mails é do n8n.
#
# Contrato (espelha automacoes/clientes/cadastro_novo.executar):
#   entrada: job = {
#       "tipo": 1,                       # informativo
#       "modo_fantasma": True,
#       "envios": [
#           {"lab_id": "bayer", "periodo_ini": "01/05/2026", "periodo_fim": "31/05/2026"},
#           ...
#       ]
#   }
#   saída: {
#       "status": "Sucesso" | "Sucesso com avisos" | "Erro",
#       "msg": "...",
#       "pasta": "/abs/caminho/da/execucao",
#       "itens": [{"lab_id": "bayer", "nome": "Bayer", "arquivos": ["estoque_privado.xlsx", "venda_*.xlsx"]}],
#       "avisos": [...]
#   }
#
# Otimização (dedup): cada combinação única é gerada UMA vez —
#   • estoque por setor (privado/público);
#   • venda por (modelo + operação + período).
# Assim o estoque privado sai 1x (todos reusam) e a venda Sandoz 1x.
# ══════════════════════════════════════════════════════════════
import os
import re
import sys
import json
import time
import shutil
import traceback
from datetime import datetime

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# Login/navegador reaproveitados do robô de cadastro (não reescrevemos login).
from automacoes.clientes.navegacao_erp import iniciar_navegador, fazer_login
from automacoes.relatorios.labs_config import ESTOQUE, DEMANDA, LABS, SETOR_LABEL, OP_LABEL
from automacoes.relatorios import pos_processamento

# Pasta-base onde cada execução cria sua subpasta de downloads.
PASTA_BASE = os.getenv("RELATORIOS_DOWNLOAD_DIR", os.path.join(os.getcwd(), "downloads_relatorios"))
TIMEOUT_PADRAO = 30
TIMEOUT_DOWNLOAD = 180

# Fator de lentidão p/ observar o robô. 1.0 = normal; 3.0 = bem lento.
# Vem do env RELATORIOS_SLOWMO ou do job ("modo_lento": true → 3.0).
try:
    _FATOR_LENTIDAO = float(os.getenv("RELATORIOS_SLOWMO", "1") or "1")
except ValueError:
    _FATOR_LENTIDAO = 1.0


def _sleep(segundos):
    """Pausa multiplicada pelo fator de lentidão (modo observação)."""
    time.sleep(segundos * _FATOR_LENTIDAO)


# ─────────────────────────── Helpers de UI ───────────────────────────
def _js_click(driver, elemento):
    """Clique forçado via JS — padrão do projeto p/ não falhar no NL."""
    driver.execute_script("arguments[0].click();", elemento)


def _set_valor(driver, item_id, valor):
    """Define o valor de um item APEX (<select>/<input>) de forma robusta.

    Usa a API do próprio APEX — apex.item(id).setValue() — que inicializa o
    widget e dispara as cascatas/refresh corretamente, sem depender de clique
    nem do item já estar focado (o dispatchEvent "cru" só pega depois que o
    widget foi ativado). Faz fallback para value+events e, em <select>, para
    o Select nativo do Selenium.
    """
    el = WebDriverWait(driver, TIMEOUT_PADRAO).until(
        EC.presence_of_element_located((By.ID, item_id))
    )
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
    ok = driver.execute_script(
        """
        var id = arguments[0], val = arguments[1];
        var el = document.getElementById(id);
        if (!el) { return false; }
        // Ativa o widget como um usuário (foco + mouse) — o APEX só anexa
        // alguns listeners no 1º foco; sem isso o change "não pega". Foi o
        // que destravou no teste manual. Tudo via DOM = headless-safe (não
        // abre o dropdown nativo do SO).
        try { el.focus(); } catch (e) {}
        ['mousedown', 'mouseup', 'click'].forEach(function (t) {
            el.dispatchEvent(new MouseEvent(t, {bubbles: true}));
        });
        if (window.apex && apex.item) {            // caminho oficial do APEX
            try { apex.item(id).setValue(val); } catch (e) {}
        }
        el.value = val;                            // garante o value no DOM
        el.dispatchEvent(new Event('input',  {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        if (window.jQuery) {                       // APEX escuta via jQuery
            try { jQuery(el).trigger('change'); } catch (e) {}
        }
        try { el.blur(); } catch (e) {}
        return el.value === String(val);
        """,
        item_id, valor,
    )
    if not ok and el.tag_name.lower() == "select":  # rede de segurança
        try:
            Select(el).select_by_value(str(valor))
        except Exception:  # noqa: BLE001
            pass
    return el


def _definir_select(driver, item_id, valor_alvo, rotulo="Filtro"):
    """Garante que um <select> termine no VALOR ALVO, forçando uma transição REAL.

    Usado tanto na Operação (Público/Privado) da Demanda quanto no Setor
    (Público/Privado) do Estoque. O robô "nunca se perde": lê o que está
    selecionado e força a transição para o alvo. Se já estiver no alvo, dá um
    'wiggle' (seleciona a outra opção e volta) para garantir que o change/refresh
    da IR dispare — caso contrário, reaplicar o mesmo valor poderia não disparar
    nada (era o que deixava o setor "preso" no anterior). Verifica o resultado no fim.

    `rotulo` é só cosmético (aparece no log): "Operação" na Demanda, "Setor" no Estoque.
    """
    valor_alvo = str(valor_alvo)
    el = WebDriverWait(driver, TIMEOUT_PADRAO).until(
        EC.presence_of_element_located((By.ID, item_id))
    )
    atual = Select(el).first_selected_option.get_attribute("value")
    if atual == valor_alvo:
        # Já está no alvo → "wiggle" p/ forçar uma troca real (vai na outra e volta).
        outra = next(
            (o.get_attribute("value") for o in Select(el).options
             if o.get_attribute("value") and o.get_attribute("value") != valor_alvo),
            None,
        )
        if outra:
            _set_valor(driver, item_id, outra)
            _sleep(0.4)
    _set_valor(driver, item_id, valor_alvo)
    _sleep(0.3)
    final = Select(driver.find_element(By.ID, item_id)).first_selected_option.get_attribute("value")
    print(f"> 🔀 {rotulo}: alvo {valor_alvo} (estava {atual}) → selecionada {final}")
    return final


def _selecionar_modelo(driver, item_id, value=None, label=None):
    """Troca o 'saved report' (modelo) por VALUE (preferido) ou por NOME (label).

    Por label: procura a <option> cujo texto contenha o trecho informado e
    resolve o value real em tempo de execução — útil quando o relatório foi
    criado depois (não temos o value) ou quando o número visível muda.

    A APLICAÇÃO usa o _set_valor robusto (apex.item + jQuery 'change' +
    ativação por mouse). É ESSENCIAL: trocar o relatório salvo é o que define
    o laboratório; com o dispatchEvent cru o APEX não recarregava a IR e o
    relatório ficava preso no anterior.
    """
    el = WebDriverWait(driver, TIMEOUT_PADRAO).until(
        EC.presence_of_element_located((By.ID, item_id))
    )
    if not value and label:
        alvo = next(
            (op for op in Select(el).options if label.lower() in op.text.lower()),
            None,
        )
        if alvo is None:
            raise ValueError(f"Modelo '{label}' não encontrado no select {item_id}.")
        value = alvo.get_attribute("value")
    _set_valor(driver, item_id, value)
    return value


class ErroRedirecionamentoNL(Exception):
    """O NL abriu a página de erro do navegador (ERR_TOO_MANY_REDIRECTS /
    'Redirecionamento em excesso') no lugar da tela — o iframe do APEX não
    carregou. É intermitente (costuma ocorrer na 1ª execução do dia); o robô
    trata refazendo o login e tentando de novo (ver _abrir_favorito_resiliente)."""


def _eh_pagina_erro_chrome(driver):
    """True se o CONTEXTO ATUAL é a página de erro do Chrome de redirecionamento
    em excesso (a do print: ícone de documento quebrado + 'Redirecionamento em
    excesso'). Casa pelo id interno do Chrome e, por segurança, pelo texto."""
    try:
        if (driver.find_elements(By.ID, "sub-frame-error")
                or driver.find_elements(By.ID, "main-frame-error")):
            return True
        corpo = driver.find_elements(By.TAG_NAME, "body")
        txt = (corpo[0].text if corpo else "").lower()
        return ("redirecionamento em excesso" in txt
                or "err_too_many_redirects" in txt
                or "too many redirects" in txt)
    except Exception:  # noqa: BLE001
        return False


def _entrar_na_tela(driver, marker_id, timeout=TIMEOUT_PADRAO):
    """Posiciona o Selenium no contexto onde o marcador da tela existe.

    Portais que hospedam apps Oracle APEX costumam renderizar a tela dentro
    de um <iframe> (e os diálogos modais do APEX TAMBÉM são iframes). O
    Selenium só "enxerga" o que está no contexto atual, então aqui:
      1) se abriu nova aba/janela, vai para a mais recente;
      2) procura o marcador no documento principal;
      3) se não achar, entra em cada <iframe> (1 nível) até encontrá-lo.
    Deixa o driver JÁ posicionado no contexto certo e diz onde achou.

    Se em vez da tela o NL abrir a página de erro do navegador
    (ERR_TOO_MANY_REDIRECTS), levanta ErroRedirecionamentoNL NA HORA — sem
    esperar o timeout inteiro — para o retry entrar em ação rápido.
    """
    fim = time.time() + timeout
    while time.time() < fim:
        if len(driver.window_handles) > 1:                  # 1) nova janela/aba
            driver.switch_to.window(driver.window_handles[-1])
        driver.switch_to.default_content()                  # 2) documento principal
        if driver.find_elements(By.ID, marker_id):
            return "topo"
        if _eh_pagina_erro_chrome(driver):                  # topo virou página de erro
            raise ErroRedirecionamentoNL(
                "O NL abriu a página de erro do navegador (redirecionamento em "
                "excesso) no lugar da tela.")
        for fr in driver.find_elements(By.TAG_NAME, "iframe"):  # 3) iframes
            driver.switch_to.default_content()
            try:
                driver.switch_to.frame(fr)
            except Exception:  # noqa: BLE001
                continue
            if driver.find_elements(By.ID, marker_id):
                return "iframe"
            if _eh_pagina_erro_chrome(driver):              # iframe virou página de erro
                driver.switch_to.default_content()
                raise ErroRedirecionamentoNL(
                    "O NL abriu a página de erro do navegador (redirecionamento em "
                    "excesso) no lugar da tela.")
        driver.switch_to.default_content()
        _sleep(0.5)
    raise TimeoutException(
        f"Marcador '{marker_id}' não encontrado no documento principal nem em iframes."
    )


def _abrir_favorito(driver, invoker_code, marker_id):
    """Abre a estrela de favoritos, clica no atalho pelo código invoker
    (estável) e entra no contexto (iframe/aba) onde a tela carregou."""
    wait = WebDriverWait(driver, TIMEOUT_PADRAO)
    driver.switch_to.default_content()
    estrela = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//a[contains(@onclick, 'tab-header-menu-favoritos')]")))
    _js_click(driver, estrela)
    _sleep(1.5)
    # Localiza o <a> do favorito pelo trecho do href: invoker(1,'N&L@....')
    link = wait.until(EC.presence_of_element_located(
        (By.XPATH, f"//a[contains(@href, \"{invoker_code}\")]")))
    _js_click(driver, link)
    _sleep(2.5)
    onde = _entrar_na_tela(driver, marker_id)
    print(f"> 🧭 Tela carregada (marcador em: {onde}).")


def _recuperar_sessao_nl(driver):
    """Recuperação após o NL abrir a página de erro: refaz o login (re-navega
    do zero e reautentica), o que limpa o loop de redirecionamento por sessão
    velha. Best-effort — devolve True se o login voltou a valer."""
    try:
        ok = fazer_login(driver)
    except Exception as e:  # noqa: BLE001
        print(f"> ⚠️ Recuperação: erro ao refazer login: {type(e).__name__}: {e}")
        return False
    if not ok:
        print("> ⚠️ Recuperação: re-login não confirmou o carregamento do menu.")
    return ok


def _abrir_favorito_resiliente(driver, invoker_code, marker_id, tentativas=3):
    """Abre o favorito TOLERANDO o erro intermitente do NL de 'Redirecionamento
    em excesso' (ERR_TOO_MANY_REDIRECTS), comum na 1ª execução do dia.

    Em caso de falha (página de erro do navegador OU a tela não carregou),
    refaz o login e tenta de novo, até `tentativas` vezes. Só desiste — e aí
    levanta o erro, que vira aviso ao TI — depois de esgotar as tentativas."""
    ultimo = None
    for tentativa in range(1, tentativas + 1):
        try:
            _abrir_favorito(driver, invoker_code, marker_id)
            if tentativa > 1:
                print(f"> ✅ NL recuperado na tentativa {tentativa}/{tentativas}.")
            return
        except (ErroRedirecionamentoNL, TimeoutException) as e:
            ultimo = e
            motivo = ("redirecionamento em excesso"
                      if isinstance(e, ErroRedirecionamentoNL) else "a tela não carregou")
            print(f"> ⚠️ NL falhou ao abrir a tela ({motivo}) — "
                  f"tentativa {tentativa}/{tentativas}.")
            if tentativa < tentativas:
                print("> 🔄 Recuperando (refazendo login) e tentando novamente...")
                _recuperar_sessao_nl(driver)
                _sleep(3)
    # Esgotou as tentativas → erro claro (mantém o tipo p/ a tradução no aviso).
    resumo = (f"O NL continuou com erro de redirecionamento ao abrir a tela após "
              f"{tentativas} tentativas (marcador '{marker_id}').")
    if isinstance(ultimo, ErroRedirecionamentoNL):
        raise ErroRedirecionamentoNL(resumo) from ultimo
    raise TimeoutException(
        f"A tela do NL não carregou após {tentativas} tentativas "
        f"(marcador '{marker_id}').") from ultimo


def _esperar_apex_pronto(driver, timeout=TIMEOUT_PADRAO):
    """Best-effort: espera o overlay de processamento do APEX sumir."""
    fim = time.time() + timeout
    seletor = ".u-Processing, .a-IRR-loading, #apex_wait_overlay, .apex_wait"
    while time.time() < fim:
        try:
            visiveis = [e for e in driver.find_elements(By.CSS_SELECTOR, seletor) if e.is_displayed()]
        except Exception:
            visiveis = []
        if not visiveis:
            return
        _sleep(0.5)


def _clicar_real(driver, el):
    """Clique robusto p/ widgets APEX (a-Menu, diálogos): dispara a sequência
    de eventos de mouse via DOM. Um .click() simples às vezes não aciona o
    handler do menu/itens do Interactive Report."""
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
    driver.execute_script(
        "var el = arguments[0];"
        "['mouseover','mousedown','mouseup','click'].forEach(function (t) {"
        "  el.dispatchEvent(new MouseEvent(t, {bubbles: true, cancelable: true, view: window}));"
        "});",
        el,
    )


def _clicar_item_menu(driver, menu_el, texto, id_fallback=None):
    """Clica num item do menu Ações pelo TEXTO visível (o id numérico do APEX,
    ex.: '..._menu_14i', muda quando o menu é reconstruído). Cai para o id
    informado se o texto não bater."""
    alvo = None
    for botao in menu_el.find_elements(By.CSS_SELECTOR, "button.a-Menu-label"):
        try:
            if botao.text.strip().lower() == texto.lower():
                alvo = botao
                break
        except Exception:  # noqa: BLE001
            continue
    if alvo is None and id_fallback:
        alvo = driver.find_element(By.ID, id_fallback)
    if alvo is None:
        raise TimeoutException(f"Item '{texto}' não encontrado no menu de Ações.")
    _clicar_real(driver, alvo)


def _fechar_dialogo(driver):
    """Fecha o diálogo de download (jQuery UI) clicando no 'X' do titlebar.

    Best-effort: o diálogo precisa sumir para não atrapalhar o próximo
    download. Não falha se não houver diálogo aberto."""
    try:
        for botao in driver.find_elements(By.CSS_SELECTOR, "button.ui-dialog-titlebar-close"):
            if botao.is_displayed():
                _clicar_real(driver, botao)
                _sleep(0.4)
                return True
    except Exception:  # noqa: BLE001
        pass
    return False


def _baixar_csv(driver, cfg, pasta, destino_basename):
    """Abre Ações → Download → CSV, espera baixar e converte para .xlsx.

    Retorna o caminho do .xlsx final (com nome determinístico).
    """
    wait = WebDriverWait(driver, TIMEOUT_PADRAO)
    antes = set(os.listdir(pasta))

    # 1) Abre o menu Ações e espera ele REALMENTE aparecer (display: block).
    botao_acoes = wait.until(EC.presence_of_element_located((By.ID, cfg["actions_button"])))
    _js_click(driver, botao_acoes)
    menu = wait.until(EC.visibility_of_element_located((By.ID, cfg["actions_menu"])))
    _sleep(0.3)

    # 2) Download — pelo texto (id numérico muda), com sequência de mouse.
    _clicar_item_menu(driver, menu, "Download", cfg.get("download_menu_item"))
    _sleep(0.6)

    # 3) Formato CSV no diálogo de download.
    link_csv = wait.until(EC.visibility_of_element_located((By.ID, cfg["download_csv"])))
    _clicar_real(driver, link_csv)

    # 4) Espera o .csv baixar e SÓ ENTÃO fecha o diálogo (X), para o clique
    #    no 'X' nunca interferir no download em andamento.
    caminho_csv = _esperar_download(pasta, antes)
    _fechar_dialogo(driver)

    # Relatório SEM dados: o APEX baixa um CSV de **0 byte** (tela "Item não
    # encontrado"). Não vira anexo — devolve None p/ o orquestrador pular com aviso.
    try:
        if os.path.getsize(caminho_csv) == 0:
            print(f"> ⚠️ Download vazio (0 byte) — relatório sem dados. Ignorando '{destino_basename}'.")
            os.remove(caminho_csv)
            return None
    except OSError:
        pass

    destino_xlsx = os.path.join(pasta, destino_basename + ".xlsx")
    _csv_para_xlsx(caminho_csv, destino_xlsx)
    try:
        os.remove(caminho_csv)  # mantém só o .xlsx final
    except OSError:
        pass
    print(f"> 💾 Gerado: {os.path.basename(destino_xlsx)}")
    return destino_xlsx


def _esperar_download(pasta, arquivos_antes, timeout=TIMEOUT_DOWNLOAD):
    """Espera surgir o ARQUIVO DE RESULTADO (.csv) novo e estável.

    Considera só extensões de resultado e ignora artefatos transitórios do
    Chrome (downloads.htm, .tmp, .crdownload). É à prova de corrida: se o
    arquivo candidato sumir/renomear entre uma medição e outra, apenas
    continua tentando em vez de estourar FileNotFoundError.

    Aceita arquivo **de 0 byte** desde que ESTÁVEL e sem `.crdownload` — é o que
    o APEX baixa quando o relatório não tem dados ("Item não encontrado"). Sem
    isso o robô esperava esse 0 byte virar >0 pra sempre, até o timeout. Quem
    decide o que fazer com o vazio é o `_baixar_csv` (devolve None).
    """
    EXT_RESULTADO = (".csv", ".xlsx", ".xls")
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            atuais = set(os.listdir(pasta))
        except OSError:
            atuais = set()
        baixando = any(a.endswith(".crdownload") for a in atuais)
        novos = [
            a for a in (atuais - arquivos_antes)
            if a.lower().endswith(EXT_RESULTADO)
        ]
        if novos and not baixando:
            caminho = os.path.join(pasta, novos[0])
            try:
                tamanho_1 = os.path.getsize(caminho)
                _sleep(1.2)
                ainda_baixando = any(x.endswith(".crdownload") for x in os.listdir(pasta))
                # Estável + sem .crdownload = download concluído. Aceita 0 byte
                # (relatório vazio) — o _baixar_csv trata o vazio depois.
                if not ainda_baixando and os.path.getsize(caminho) == tamanho_1:
                    return caminho
            except OSError:
                pass  # candidato sumiu/renomeou — segue tentando
        _sleep(0.5)
    raise TimeoutError("Download (.csv) não foi concluído dentro do tempo limite.")


def _csv_para_xlsx(caminho_csv, destino_xlsx):
    """Converte o CSV do APEX em .xlsx (detecta separador/encoding)."""
    import pandas as pd

    df = None
    ultimo_erro = None
    for enc in ("utf-8-sig", "latin-1"):
        try:
            df = pd.read_csv(
                caminho_csv, sep=None, engine="python",
                dtype=str, encoding=enc, keep_default_na=False,
            )
            break
        except Exception as e:  # noqa: BLE001
            ultimo_erro = e
            df = None
    if df is None:
        raise ValueError(f"Falha ao ler o CSV baixado: {ultimo_erro}")
    df.to_excel(destino_xlsx, index=False)
    return destino_xlsx


def _estilizar_xlsx(caminho, titulo="", subtitulo=""):
    """Aplica o estilo-padrão Ciamed ao .xlsx, in-place (best-effort):
    faixa-título CIAMED + subtítulo, cabeçalho teal, congelamento, filtro,
    larguras automáticas, linhas zebradas e números alinhados à direita.
    Os VALORES não são alterados (seguem texto pt-BR) — só a aparência."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    TEAL, TEAL_HDR, ZEBRA, BRANCO, SUB = "00B3B2", "0A9D9C", "F4F6F8", "FFFFFF", "E6F7F7"
    re_num = re.compile(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$|^-?\d+(,\d+)?$")

    wb = load_workbook(caminho)
    ws = wb.active
    ncols, nrows = ws.max_column, ws.max_row
    if ncols == 0 or nrows == 0:
        return

    # Larguras pelo conteúdo (mín. 10, máx. 60) — inclui o header (linha 1).
    larguras = {}
    for col in range(1, ncols + 1):
        m = 0
        for row in range(1, nrows + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                m = max(m, len(str(v)))
        larguras[col] = min(max(m + 5, 12), 65)  # +respiro nas colunas

    # Colunas numéricas (amostra das 1ªs linhas de dados) → alinhar à direita.
    col_num = {}
    for col in range(1, ncols + 1):
        amostra = num = 0
        for row in range(2, min(nrows, 41) + 1):
            s = (ws.cell(row=row, column=col).value or "")
            s = str(s).strip()
            if not s:
                continue
            amostra += 1
            num += 1 if re_num.match(s) else 0
        col_num[col] = amostra > 0 and num / amostra >= 0.8

    # Faixa-título (2 linhas no topo).
    ws.insert_rows(1, amount=2)
    fill_teal = PatternFill("solid", fgColor=TEAL)
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).fill = fill_teal
        ws.cell(row=2, column=col).fill = fill_teal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    t1 = ws.cell(row=1, column=1)
    t1.value = "CIAMED"
    t1.font = Font(name="Calibri", bold=True, size=18, color=BRANCO)
    t1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    t2 = ws.cell(row=2, column=1)
    t2.value = " · ".join(x for x in (titulo, subtitulo) if x)
    t2.font = Font(name="Calibri", size=10, color=SUB)
    t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18

    # Cabeçalho da tabela (agora na linha 3).
    HDR = 3
    fill_hdr = PatternFill("solid", fgColor=TEAL_HDR)
    font_hdr = Font(name="Calibri", bold=True, color=BRANCO)
    al_l = Alignment(horizontal="left", vertical="center")
    al_r = Alignment(horizontal="right", vertical="center")
    al_c = Alignment(horizontal="center", vertical="center")
    for col in range(1, ncols + 1):
        c = ws.cell(row=HDR, column=col)
        c.fill, c.font = fill_hdr, font_hdr
        c.alignment = al_c  # títulos das colunas centralizados
    ws.row_dimensions[HDR].height = 20

    # Corpo: zebra + alinhamento.
    fill_zebra = PatternFill("solid", fgColor=ZEBRA)
    font_body = Font(name="Calibri", size=11)
    inicio, ult = HDR + 1, ws.max_row
    for row in range(inicio, ult + 1):
        zebra = (row - inicio) % 2 == 1
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            if zebra:
                c.fill = fill_zebra
            c.font = font_body
            c.alignment = al_r if col_num[col] else al_l

    # Larguras, filtro e congelamento (mantém faixa + cabeçalho fixos ao rolar).
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A{HDR}:{get_column_letter(ncols)}{ult}"
    ws.freeze_panes = f"A{inicio}"

    wb.save(caminho)


# ─────────────────────── Geração de cada relatório ───────────────────────
def gerar_analise_estoque(driver, setor, modelo_value, modelo_label, nome_base, pasta):
    """Tela A — gera a Análise de Estoque para (modelo, setor).

    Ordem CRÍTICA: **modelo → Consultar → setor → Consultar → baixar**. Selecionar
    o saved report (modelo) RESETA o filtro P366_FILTRO para branco; por isso o
    setor é definido DEPOIS do modelo (e de um Consultar), senão o relatório vinha
    com os dois setores misturados. `nome_base` é o nome amigável do arquivo, já
    montado pelo orquestrador (ex.: "Estoque - Bayer - Privado").
    """
    rotulo = modelo_value or modelo_label
    print(f"> 📦 Análise de Estoque — modelo {rotulo} | setor: {setor}")
    wait = WebDriverWait(driver, TIMEOUT_PADRAO)
    wait.until(EC.presence_of_element_located((By.ID, ESTOQUE["item_marcador"])))

    # 1) Modelo PRIMEIRO (selecionar o saved report zera o filtro de setor).
    _selecionar_modelo(driver, ESTOQUE["saved_reports"], value=modelo_value, label=modelo_label)
    _esperar_apex_pronto(driver)
    _sleep(0.8)

    # Diagnóstico: nome do relatório salvo ativo.
    try:
        sel = Select(driver.find_element(By.ID, ESTOQUE["saved_reports"]))
        print(f"> 🔖 Relatório salvo ativo: '{sel.first_selected_option.text.strip()}'"
              f" (value={sel.first_selected_option.get_attribute('value')})")
    except Exception as e:  # noqa: BLE001
        print(f"> ⚠️ Não consegui ler o relatório salvo ativo: {e}")

    # 2) Consultar com o modelo — deixa a IR assentar ANTES de mexer no setor
    #    (o reset do filtro pela troca de modelo já aconteceu neste ponto).
    _js_click(driver, driver.find_element(By.ID, ESTOQUE["btn_consultar"]))
    _esperar_apex_pronto(driver)
    _sleep(1.0)

    # 3) SÓ AGORA o setor (Público/Privado) — depois do modelo, p/ não ser resetado.
    _definir_select(driver, ESTOQUE["item_filtro"], ESTOQUE["filtro_setor"][setor], rotulo="Setor")
    _sleep(0.5)

    # 4) Consultar de novo, aplicando o setor.
    _js_click(driver, driver.find_element(By.ID, ESTOQUE["btn_consultar"]))
    _esperar_apex_pronto(driver)
    _sleep(1.5)

    # Se o relatório vier vazio (ex.: Sun Ilumya sem estoque Público num mês), o
    # _baixar_csv detecta o CSV de 0 byte e devolve None → o orquestrador pula com aviso.
    return _baixar_csv(driver, ESTOQUE, pasta, nome_base)


def gerar_demanda(driver, modelo_value, modelo_label, operacao, data_ini, data_fim, nome_base, pasta):
    """Tela B — gera a Demanda Fornecedor para (modelo, operação, período).

    `nome_base` é o nome amigável do arquivo, montado pelo orquestrador
    (ex.: "Mapa de Venda - Roche - 01-05-2026 a 31-05-2026").
    """
    rotulo = modelo_value or modelo_label
    print(f"> 🧾 Demanda Fornecedor — modelo {rotulo} | op {operacao} | {data_ini}→{data_fim}")
    wait = WebDriverWait(driver, TIMEOUT_PADRAO)
    wait.until(EC.presence_of_element_located((By.ID, DEMANDA["item_marcador"])))

    # 1) PRIMEIRO as datas (regra do lab); DEPOIS a operação (com transição
    #    real até o alvo, para Público/Privado nunca ficar "preso").
    _set_valor(driver, DEMANDA["data_ini"], data_ini)
    _set_valor(driver, DEMANDA["data_fim"], data_fim)
    _definir_select(driver, DEMANDA["operacao"], operacao, rotulo="Operação")
    _sleep(0.5)

    # 2) DEPOIS troca o relatório salvo (modelo) — última ação antes do
    #    Consultar — e espera a IR recarregar.
    _selecionar_modelo(driver, DEMANDA["saved_reports"], value=modelo_value, label=modelo_label)
    _esperar_apex_pronto(driver)
    _sleep(1.0)

    # Diagnóstico + captura o NOME do relatório salvo ativo (vira o nome do arquivo).
    modelo_nome = None
    try:
        sel = Select(driver.find_element(By.ID, DEMANDA["saved_reports"]))
        modelo_nome = sel.first_selected_option.text.strip()
        print(f"> 🔖 Relatório salvo ativo agora: '{modelo_nome}'"
              f" (value={sel.first_selected_option.get_attribute('value')})")
    except Exception as e:  # noqa: BLE001
        print(f"> ⚠️ Não consegui ler o relatório salvo ativo: {e}")

    # 3) SÓ ENTÃO Consulta.
    _js_click(driver, driver.find_element(By.ID, DEMANDA["btn_consultar"]))
    _esperar_apex_pronto(driver)
    _sleep(1.5)

    # Venda vazia (lab sem vendas no período): _baixar_csv detecta o CSV de 0 byte
    # e devolve None → o orquestrador pula com aviso. Retorna (caminho, nome_modelo).
    return _baixar_csv(driver, DEMANDA, pasta, nome_base), modelo_nome


def aplicar_pos_processamento(caminho_venda, lab):
    """Aplica os hooks do lab SOBRE o próprio arquivo de venda (in-place),
    preservando o nome do relatório. Escreve num temporário e troca o original
    (evita corromper o arquivo se o hook falhar no meio)."""
    for nome_hook in lab.get("pos", []):
        func = pos_processamento.HOOKS.get(nome_hook)
        if not func:
            print(f"> ⚠️ [pos] hook desconhecido: {nome_hook} (ignorado).")
            continue
        raiz, ext = os.path.splitext(caminho_venda)
        tmp = f"{raiz}.__pos_{nome_hook}{ext}"
        func(caminho_venda, tmp)
        os.replace(tmp, caminho_venda)
    return caminho_venda


# ─────────────────────── Tradução de erros (avisos) ───────────────────────
def _explicar_erro(exc):
    """Traduz a exceção em (explicação, sugestão) no português, para o e-mail de
    aviso. Casa primeiro pela MENSAGEM (mais específico) e depois pelo TIPO."""
    nome = type(exc).__name__
    msg = (str(exc) or "").lower()

    # 1) Por mensagem (casos mais específicos primeiro)
    if ("redirecionamento em excesso" in msg or "err_too_many_redirects" in msg
            or "too many redirects" in msg or "página de erro do navegador" in msg):
        return ("O NL abriu a página de erro do navegador (redirecionamento em excesso) "
                "ao carregar a tela — costuma acontecer na 1ª execução do dia. O robô "
                "refez o login e tentou de novo algumas vezes, mas o erro persistiu.",
                "Rode novamente em alguns minutos. Se continuar, a TI deve verificar o NL "
                "(nlwebprod.ciamed.com.br) — provável loop de redirecionamento/sessão.")
    if "only supports chrome version" in msg or "this version of chromedriver" in msg:
        return ("A versão do Chrome e a do ChromeDriver estão incompatíveis.",
                "TI: atualizar o ChromeDriver para a mesma versão do Chrome da máquina do robô.")
    if "chrome not reachable" in msg or "cannot connect to chrome" in msg:
        return ("O robô não conseguiu falar com o Chrome (o navegador não respondeu/não abriu).",
                "Rode novamente. Se persistir, reinicie a máquina do robô e confira a instalação do Chrome.")
    if "disconnected" in msg or "session deleted" in msg:
        return ("A sessão do navegador caiu durante a execução.",
                "Rode novamente. Se persistir, reinicie a máquina do robô.")
    if "net::err" in msg or "timed out receiving message from renderer" in msg:
        return ("Houve uma falha de rede/carregamento ao acessar o NL.",
                "Verifique a conexão/VPN com o NL e rode novamente.")

    # 2) Por tipo da exceção
    DIC = {
        "NoSuchWindowException": (
            "A janela do Chrome foi fechada durante a execução — o robô perdeu o controle do navegador.",
            "Rode novamente. No modo visível, não feche a janela do Chrome até o robô terminar."),
        "InvalidSessionIdException": (
            "A sessão do navegador foi perdida no meio da execução.",
            "Rode novamente. Se persistir, reinicie a máquina do robô."),
        "SessionNotCreatedException": (
            "Não foi possível iniciar o navegador (Chrome/ChromeDriver).",
            "TI: conferir a compatibilidade entre o Chrome e o ChromeDriver."),
        "WebDriverException": (
            "Falha de comunicação com o navegador (Chrome/ChromeDriver).",
            "Rode novamente. Se persistir, a TI deve conferir o Chrome/ChromeDriver."),
        "TimeoutException": (
            "Uma tela do NL demorou demais para responder (lentidão/queda do NL, VPN, ou um elemento não apareceu).",
            "Tente novamente em alguns minutos. Se travar sempre na mesma etapa, o NL pode ter mudado de tela — avisar o desenvolvedor."),
        "NoSuchElementException": (
            "Um campo ou botão esperado não foi encontrado na tela do NL.",
            "Tente novamente. Se persistir, o layout do NL provavelmente mudou — avisar o desenvolvedor."),
        "ElementClickInterceptedException": (
            "Algo ficou na frente do elemento que o robô tentou clicar (pop-up/sobreposição do NL).",
            "Tente novamente."),
        "ElementNotInteractableException": (
            "Um elemento ainda não estava pronto (visível/clicável) quando o robô tentou usá-lo.",
            "Tente novamente."),
        "StaleElementReferenceException": (
            "A página recarregou e o robô perdeu a referência de um elemento.",
            "Tente novamente."),
        "UnexpectedAlertPresentException": (
            "Apareceu um alerta do navegador que interrompeu o robô.",
            "Tente novamente. Se persistir, avisar o desenvolvedor."),
        "JavascriptException": (
            "Houve um erro ao executar um comando interno na página do NL.",
            "Tente novamente. Se persistir, avisar o desenvolvedor."),
        "FileNotFoundError": (
            "O arquivo baixado do NL não foi encontrado (o download falhou ou não concluiu).",
            "Tente novamente. Se persistir, verifique a conexão com o NL e a pasta de downloads."),
        "TimeoutError": (
            "O download de um relatório não concluiu no tempo esperado.",
            "Tente novamente em alguns minutos (pode ser lentidão do NL)."),
        "PermissionError": (
            "Um arquivo de planilha estava bloqueado (provavelmente aberto no Excel).",
            "Feche a planilha aberta e rode novamente."),
        "ValueError": (
            "Um arquivo baixado veio vazio ou em formato inesperado.",
            "Tente novamente. Se persistir, avisar o desenvolvedor."),
        "KeyError": (
            "Faltou uma informação esperada na configuração de um laboratório.",
            "Avisar o desenvolvedor (possível erro de configuração de lab)."),
    }
    if nome in DIC:
        return DIC[nome]
    return ("Ocorreu um erro inesperado durante a execução do robô.",
            "Rode novamente. Se persistir, encaminhe o detalhe técnico ao desenvolvedor.")


def _erro_retriavel(exc):
    """Diz se vale a pena o n8n RE-EXECUTAR o robô inteiro (Chrome novo) para
    este erro. `True` p/ falhas TRANSITÓRIAS (NL/rede/sessão/redirecionamento,
    lentidão, navegador caído) — uma nova execução do zero costuma resolver.
    `False` p/ erros que uma re-execução NÃO conserta (config de lab, arquivo
    de planilha travado no Excel, dado em formato inesperado)."""
    if type(exc).__name__ in ("KeyError", "PermissionError", "ValueError"):
        return False
    return True


def _resultado_erro(ctx, pasta, avisos, exc=None, explicacao=None, sugestao=None,
                    detalhe=None, msg=None, retriavel=None):
    """Monta o resultado de erro ESTRUTURADO que o n8n usa para o e-mail de aviso:
    o quê (explicação), onde (etapa/lab/período/progresso) e o detalhe técnico.
    `retriavel` diz ao n8n se ele deve re-executar o robô (rede de segurança)."""
    if exc is not None:
        nome = type(exc).__name__
        det = (str(exc) or "").strip().splitlines()[0] if str(exc).strip() else "(sem mensagem)"
        detalhe = f"{nome}: {det}"
        explicacao, sugestao = _explicar_erro(exc)
        msg = detalhe
        if retriavel is None:
            retriavel = _erro_retriavel(exc)
    return {
        "status": "Erro",
        "msg": msg or detalhe or "Erro na execução.",
        "etapa": ctx.get("etapa", ""),
        "lab": ctx.get("lab", ""),
        "periodo": ctx.get("periodo", ""),
        "progresso": ctx.get("progresso", ""),
        "explicacao": explicacao or "Ocorreu um erro inesperado durante a execução do robô.",
        "sugestao": sugestao or "Rode novamente; se persistir, avise o desenvolvedor.",
        "detalhe_tecnico": detalhe or msg or "",
        # Rede de segurança do n8n: só re-executa o robô se o erro for transitório.
        "retriavel": bool(retriavel) if retriavel is not None else False,
        "pasta": pasta,
        "itens": [],
        "avisos": avisos,
    }


# ─────────────────────────── Orquestração ───────────────────────────
def executar(job):
    """Ponto de entrada chamado pelo Hub (síncrono). Ver contrato no topo."""
    envios = job.get("envios", []) or []
    # Padrão FALSE p/ observar o robô durante os testes (Chrome visível).
    # ⚠️ PRODUÇÃO/servidor sem tela: volte para True (headless) ou mande
    # "modo_fantasma": true no job — senão o Chrome não sobe sem display.
    modo_fantasma = job.get("modo_fantasma", False)
    avisos = []
    # Contexto da execução: alimenta o aviso de erro com ONDE o robô parou.
    ctx = {"etapa": "Inicialização", "lab": "", "periodo": "", "progresso": ""}

    # Modo observação: deixa os cliques/pausas mais lentos para acompanhar.
    if job.get("modo_lento"):
        global _FATOR_LENTIDAO
        _FATOR_LENTIDAO = max(_FATOR_LENTIDAO, 3.0)
        print(f"> 🐢 Modo lento ligado (fator {_FATOR_LENTIDAO}x).")

    # Valida lab_ids cedo (evita abrir navegador à toa).
    desconhecidos = [e.get("lab_id") for e in envios if e.get("lab_id") not in LABS]
    if desconhecidos:
        return _resultado_erro(ctx, "", avisos,
            msg=f"lab_id desconhecido(s): {desconhecidos}",
            detalhe=f"lab_id desconhecido(s): {desconhecidos}",
            explicacao=f"O(s) laboratório(s) {desconhecidos} não existe(m) na configuração.",
            sugestao="Confira o lab_id no Disparo Manual/Agenda (precisa casar com a lista de labs).",
            retriavel=False)  # erro de config: re-executar não resolve
    if not envios:
        return _resultado_erro(ctx, "", avisos,
            msg="Nenhum envio informado no job.",
            detalhe="Nenhum envio informado no job.",
            explicacao="O robô foi acionado sem nenhum envio para processar.",
            sugestao="Verifique se a Agenda/Disparo Manual gerou algum envio válido para hoje.",
            retriavel=False)  # nada a processar: re-executar não resolve

    pasta = os.path.join(PASTA_BASE, datetime.now().strftime("exec_%Y%m%d_%H%M%S"))
    os.makedirs(pasta, exist_ok=True)
    print(f"> 🚀 Iniciando robô de relatórios — {len(envios)} envio(s). Pasta: {pasta}")

    driver = iniciar_navegador(modo_fantasma=modo_fantasma)
    # Direciona os downloads do Chrome para a pasta da execução (headless inclusive).
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": pasta})
    except Exception as e:  # noqa: BLE001
        avisos.append(f"Não foi possível fixar a pasta de download via CDP: {e}")

    try:
        ctx["etapa"] = "Login no NL"
        if not fazer_login(driver):
            return _resultado_erro(ctx, pasta, avisos,
                msg="Falha no login do NL.",
                detalhe="fazer_login(driver) retornou False.",
                explicacao="O robô não conseguiu entrar no NL (login não concluído).",
                sugestao="Confira usuário/senha do NL e se o sistema está no ar; depois rode novamente.",
                retriavel=True)  # login pode falhar por instabilidade momentânea do NL

        # ---- 1) ANÁLISE DE ESTOQUE (gera cada (modelo, setor) único 1x) ----
        # Cada lab tem o SEU modelo de estoque → a unidade é (modelo, setor),
        # espelhando o (modelo, op, período) da Demanda. Sandoz RS/SC e SP
        # compartilham o modelo → o dedup gera o arquivo 1x só.
        combos_est_alvo = set()
        for e in envios:
            lab = LABS[e["lab_id"]]
            for s in lab["estoque"]:
                combos_est_alvo.add((lab["modelo_estoque"], s))
        total_est, feito_est = len(combos_est_alvo), 0
        estoque_paths = {}  # (modelo_estoque, setor) -> caminho
        estilos = {}  # caminho -> (titulo, subtitulo) p/ a estilização final
        if combos_est_alvo:
            ctx["etapa"] = "Abrindo a tela de Análise de Estoque no NL"
            _abrir_favorito_resiliente(driver, ESTOQUE["favorito_invoker"], ESTOQUE["item_marcador"])
            for e in envios:
                lab = LABS[e["lab_id"]]
                nome_disp = lab["nome"]                   # exibição (subtítulo/manifesto)
                nome_fs = nome_disp.replace("/", "-")     # nome de arquivo (sanitiza a "/")
                for s in lab["estoque"]:
                    chave = (lab["modelo_estoque"], s)
                    if chave not in estoque_paths:
                        rotulo = SETOR_LABEL[s]           # "Privado"/"Público" (sempre no nome)
                        ctx["etapa"] = f"Análise de Estoque — {nome_disp} / {rotulo}"
                        ctx["lab"], ctx["periodo"] = nome_disp, ""
                        nome_base = f"Estoque - {nome_fs} - {rotulo}"
                        # Tolerante a relatório vazio/falha: não derruba o lote.
                        # gerar_analise_estoque devolve None quando o setor não tem
                        # dados (ex.: Sun Ilumya sem Público num mês).
                        erro = None
                        try:
                            p = gerar_analise_estoque(
                                driver, s, lab["modelo_estoque"],
                                lab.get("modelo_estoque_label"), nome_base, pasta,
                            )
                        except Exception as ex:  # noqa: BLE001
                            p, erro = None, ex
                        feito_est += 1
                        ctx["progresso"] = f"Estoque {feito_est}/{total_est}"
                        # Marca o combo como visto (mesmo None) p/ o dedup não repetir.
                        estoque_paths[chave] = p
                        if p is None:
                            if erro is not None:
                                avisos.append(f"Estoque '{nome_base}' deu erro e foi ignorado: "
                                              f"{type(erro).__name__}: {(str(erro).splitlines() or [''])[0]}")
                            else:
                                avisos.append(f"Estoque '{nome_base}' sem dados nesse setor — não anexado.")
                            continue
                        estilos[p] = ("Análise de Estoque", f"{nome_disp} · {rotulo} · posição atual")

        # ---- 2) DEMANDA FORNECEDOR (gera cada combo único 1x) ----
        # Conta os combos únicos só para o "progresso" no aviso de erro.
        combos_alvo = set()
        for e in envios:
            lab = LABS[e["lab_id"]]
            modelo_id = lab["modelo_venda"] or lab.get("modelo_label")
            for operacao in lab["operacoes"]:
                combos_alvo.add((modelo_id, operacao, e["periodo_ini"], e["periodo_fim"]))
        total_venda, feito_venda = len(combos_alvo), 0

        combos = {}  # (modelo_id, operacao, ini, fim) -> caminho (None se vazio/falhou)
        ctx["etapa"] = "Abrindo a tela de Demanda Fornecedor no NL"
        _abrir_favorito_resiliente(driver, DEMANDA["favorito_invoker"], DEMANDA["item_marcador"])
        for e in envios:
            lab = LABS[e["lab_id"]]
            modelo_id = lab["modelo_venda"] or lab.get("modelo_label")
            for operacao in lab["operacoes"]:
                chave = (modelo_id, operacao, e["periodo_ini"], e["periodo_fim"])
                if chave not in combos:
                    ctx["etapa"] = "Demanda Fornecedor (venda)"
                    ctx["lab"] = lab["nome"]
                    ctx["periodo"] = f'{e["periodo_ini"]} a {e["periodo_fim"]}'
                    # Nome amigável do arquivo de venda (igual ao estoque). Inclui o
                    # rótulo da operação só quando o lab tem as duas (Sun) — p/ distinguir.
                    nome_v = lab["nome"].replace("/", "-")   # sanitiza a "/" (Sandoz RS/SC)
                    periodo_arq = f"{e['periodo_ini'].replace('/', '-')} a {e['periodo_fim'].replace('/', '-')}"
                    if len(lab["operacoes"]) > 1:
                        nome_base_v = f"Mapa de Venda - {nome_v} - {OP_LABEL.get(operacao, operacao)} - {periodo_arq}"
                    else:
                        nome_base_v = f"Mapa de Venda - {nome_v} - {periodo_arq}"
                    # Tolerante a venda vazia/falha: não derruba o lote (igual ao estoque).
                    erro = None
                    try:
                        caminho, modelo_nome = gerar_demanda(
                            driver, lab["modelo_venda"], lab.get("modelo_label"), operacao,
                            e["periodo_ini"], e["periodo_fim"], nome_base_v, pasta,
                        )
                    except Exception as ex:  # noqa: BLE001
                        caminho, modelo_nome, erro = None, None, ex
                    combos[chave] = caminho  # marca o combo como visto (mesmo None)
                    feito_venda += 1
                    ctx["progresso"] = f"Estoque {total_est}/{total_est} · Venda {feito_venda}/{total_venda}"
                    if caminho is None:
                        periodo_txt = f"{e['periodo_ini']} a {e['periodo_fim']}"
                        if erro is not None:
                            avisos.append(f"Venda '{lab['nome']}' ({periodo_txt}) deu erro e foi ignorada: "
                                          f"{type(erro).__name__}: {(str(erro).splitlines() or [''])[0]}")
                        else:
                            avisos.append(f"Venda '{modelo_nome or lab['nome']}' ({periodo_txt}) sem dados — não anexada.")
                        continue
                    estilos[caminho] = ("Mapa de Venda",
                        f"{modelo_nome or lab['nome']} · {e['periodo_ini']} a {e['periodo_fim']}")

        # ---- 3) MANIFESTO (lab -> arquivos), com pós-processamento ----
        ctx["etapa"] = "Montagem dos arquivos / pós-processamento"
        ctx["lab"], ctx["periodo"] = "", ""
        itens = []
        pos_feito = set()  # combos já pós-processados (evita somar 2x no mesmo arquivo)
        for e in envios:
            lab = LABS[e["lab_id"]]
            modelo_id = lab["modelo_venda"] or lab.get("modelo_label")
            multi_op = len(lab["operacoes"]) > 1   # Sun Geral: distingue 11/17 no rótulo
            # arquivos = nomes de arquivo COM dados (p/ anexar/baixar no n8n).
            # com_dados / vazios = RÓTULOS amigáveis (p/ o e-mail de aviso ler bonito).
            arquivos, com_dados, vazios = [], [], []
            for s in lab["estoque"]:
                label = f"Estoque ({SETOR_LABEL[s]})"   # ex.: "Estoque (Privado)"
                caminho_est = estoque_paths.get((lab["modelo_estoque"], s))
                if caminho_est:
                    arquivos.append(os.path.basename(caminho_est))
                    com_dados.append(label)
                else:
                    vazios.append(label)
            for operacao in lab["operacoes"]:
                label = "Mapa de Venda" + (f" ({OP_LABEL.get(operacao, operacao)})" if multi_op else "")
                chave = (modelo_id, operacao, e["periodo_ini"], e["periodo_fim"])
                caminho_venda = combos.get(chave)
                if not caminho_venda:   # venda vazia/falha → não anexa, registra em vazios
                    vazios.append(label)
                    continue
                # Pós-processa o arquivo de venda IN-PLACE, uma única vez por combo.
                # Seguro: um lab com hook usa modelo exclusivo (ex.: Fresenius), então
                # nenhum outro lab reaproveita esse combo no dedup.
                if lab.get("pos") and chave not in pos_feito:
                    ctx["etapa"] = f"Pós-processamento — {lab['nome']}"
                    ctx["lab"] = lab["nome"]
                    aplicar_pos_processamento(caminho_venda, lab)
                    pos_feito.add(chave)
                arquivos.append(os.path.basename(caminho_venda))
                com_dados.append(label)
            itens.append({
                "id": e.get("id"),
                "lab_id": e["lab_id"],
                "nome": lab["nome"],
                "arquivos": arquivos,       # nomes de arquivo (anexos)
                "com_dados": com_dados,     # rótulos amigáveis dos que vieram COM dados
                "vazios": vazios,           # rótulos amigáveis dos que vieram SEM dados
            })

        # ---- 4) ESTILIZAÇÃO (passo final, DEPOIS do pós-processamento) ----
        # Padroniza todos os arquivos (faixa CIAMED, cabeçalho teal, freeze,
        # filtro, larguras, zebra). Best-effort: falha aqui não derruba o envio.
        for caminho, (titulo, subtitulo) in estilos.items():
            try:
                _estilizar_xlsx(caminho, titulo, subtitulo)
            except Exception as ex:  # noqa: BLE001
                avisos.append(f"Não consegui estilizar {os.path.basename(caminho)}: {ex}")

        status = "Sucesso com avisos" if avisos else "Sucesso"
        print("> ✅ Concluído.")
        return {"status": status, "msg": "Relatórios gerados.", "pasta": pasta,
                "itens": itens, "avisos": avisos}

    except Exception as e:  # noqa: BLE001
        tipo = type(e).__name__
        detalhe = (str(e) or "").strip().splitlines()[0] if str(e).strip() else "(sem mensagem)"
        print(f"> ❌ Erro [{tipo}] na etapa '{ctx.get('etapa')}': {detalhe}")
        traceback.print_exc()  # mostra a LINHA do robô que estourou
        return _resultado_erro(ctx, pasta, avisos, exc=e)
    finally:
        _sleep(2)
        driver.quit()


def main():
    """Execução standalone p/ teste.

    Aceita o job de 3 formas (a 1ª que existir):
      • caminho de um arquivo .json:  python -m ...gerar_relatorios job.json
      • JSON inline:                  python -m ...gerar_relatorios '{"envios":[...]}'
      • via stdin:                    Get-Content job.json | python -m ...gerar_relatorios

    Dica PowerShell: prefira o ARQUIVO. O PowerShell remove as aspas duplas
    de JSON inline, o que quebra o parse ("Expecting property name ...").
    """
    sys.stdout.reconfigure(encoding="utf-8")
    try:
        if len(sys.argv) > 1:
            arg = sys.argv[1]
            if os.path.isfile(arg):                       # caminho de arquivo .json
                with open(arg, "r", encoding="utf-8") as f:
                    job = json.load(f)
            else:                                          # JSON inline
                job = json.loads(arg)
        elif not sys.stdin.isatty():                       # JSON via stdin (pipe)
            job = json.load(sys.stdin)
        else:
            job = {}
    except Exception as e:  # noqa: BLE001
        print(f"> ❌ Job inválido: {e}")
        print("   Dica (PowerShell): use um arquivo .json em vez de JSON inline.")
        print("[FIM_DO_PROCESSO]")
        return
    resultado = executar(job)
    print("[RESULTADO_JSON] " + json.dumps(resultado, ensure_ascii=False))
    print("[FIM_DO_PROCESSO]")


if __name__ == "__main__":
    main()
