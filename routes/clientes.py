# ══════════════════════════════════════════════════════════════
# routes/clientes.py — Cadastro e Reativação de Clientes
# ══════════════════════════════════════════════════════════════
# Maior blueprint do Hub. Contém:
#   - Página de cadastro/reativação
#   - Ficha digital do cliente
#   - Recebimento de fichas (com enriquecimento CNPJ.ws)
#   - Fluxo de aprovação/reprovação (farmacêutica)
#   - Monitor de cadastros + exportação Excel
#   - Fila de cadastro (integração N8N)
#   - Endpoint de integração N8N → Robô RPA
# ══════════════════════════════════════════════════════════════

import os
import re
import json
from io import BytesIO
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import requests as req_ext
import uuid as uuid_lib

import banco_cadastros
from utils.cnpj_ws import consultar_cnpj_ws
from utils.fila import (
    cadastro_entrar, cadastro_liberar_proximo,
    fila_cadastro_status_dict, fila_cadastro_reset as _fila_cadastro_reset,
)
from utils.n8n_security import resume_url_confiavel, exigir_token_n8n
from utils.auth import permissao_required, admin_required
from utils.validacao import eh_uuid_valido
import banco_links
from utils.rastreio import iniciar_execucao_robo, finalizar_execucao_robo
from automacoes.clientes import cadastro_novo
from automacoes.clientes import cadastro_reativacao
from utils.cebas import consultar_cebas

import logging
logger = logging.getLogger(__name__)

clientes_bp = Blueprint('clientes', __name__)


# ── Páginas ───────────────────────────────────────────────────

@clientes_bp.route('/cadastro_clientes')
@permissao_required('clientes')
def cadastro_clientes():
    return render_template('cadastro_clientes.html')


@clientes_bp.route('/api/gerar_link_ficha', methods=['POST'])
@permissao_required('clientes')
def gerar_link_ficha():
    """
    Gera o link da ficha (V-05) e o registra no banco para controle.
    Só operadores logados com permissão 'clientes' geram links. O front
    monta a URL final com origin + /ficha_cliente?t=<token>.
    """
    dados = request.get_json(silent=True) or {}
    vendedor   = (dados.get('vendedor')  or '').strip()
    rep        = (dados.get('rep')       or '').strip()
    cap        = (dados.get('cap')       or '').strip()
    tipo       = (dados.get('tipo')      or '').strip()
    codigo_nl  = (dados.get('codigo_nl') or '').strip()
    cnpj_cliente = re.sub(r'\D', '', dados.get('cnpj_cliente') or '')

    if not vendedor or not rep or not cap or not tipo:
        return jsonify({'erro': 'Preencha vendedor, representante, captação e tipo.'}), 400
    if len(cnpj_cliente) != 14:
        return jsonify({'erro': 'Informe um CNPJ válido (14 dígitos) do cliente.'}), 400

    token = banco_links.criar_link(
        cnpj_cliente=cnpj_cliente,
        vendedor=vendedor, representante=rep, captacao=cap,
        tipo=tipo, codigo_nl=codigo_nl,
        gerado_por_id=session.get('usuario_id'),
        gerado_por_nome=session.get('usuario_nome', 'Sistema'),
    )
    return jsonify({'token': token})


@clientes_bp.route('/ficha_cliente')
def ficha_cliente():
    # Modo correção: o link vem do e-mail de reprovação (N8N) com
    # ?correcao=<uuid>. Continua protegido pelo conhecimento do UUID,
    # como antes — não exige token.
    if request.args.get('correcao'):
        return render_template('ficha_cliente.html', ficha_params=None, ficha_token='')

    # Nova ficha: exige um token de link válido, dentro da validade e
    # ainda não utilizado.
    token = request.args.get('t', '')
    link, motivo = banco_links.validar_link(token)
    if motivo:
        return render_template('ficha_link_invalido.html', motivo=motivo), 403

    ficha_params = {
        'vendedor':  link.vendedor or '',
        'rep':       link.representante or '',
        'cap':       link.captacao or '',
        'tipo':      link.tipo or '',
        'codigo_nl': link.codigo_nl or '',
    }
    return render_template('ficha_cliente.html', ficha_params=ficha_params, ficha_token=token)


# ══════════════════════════════════════════════════════════════
# DISPARO AO N8N (reutilizado pelo recebimento e pelo reenvio)
# ══════════════════════════════════════════════════════════════

def _disparar_para_n8n(submission_uuid, payload_str, docs_permitidos=None):
    """
    Envia (ou reenvia) uma ficha ao webhook do N8N, anexando os arquivos
    salvos em uploads_fichas/<uuid>. Se 'docs_permitidos' for informado,
    anexa só esses documentos; senão, anexa todos os da pasta.

    Retorna (ok, erro):
        ok=True  → N8N recebeu (HTTP < 400)
        ok=False → falha (N8N fora do ar, timeout, HTTP de erro); 'erro' = msg
    """
    n8n_url = os.getenv('N8N_WEBHOOK_CADASTRO')
    if not n8n_url:
        return False, 'N8N_WEBHOOK_CADASTRO não configurado.'

    arquivos_abertos = {}
    pasta = os.path.join('uploads_fichas', submission_uuid)
    if os.path.exists(pasta):
        for arquivo in os.listdir(pasta):
            nome_sem_ext = os.path.splitext(arquivo)[0]
            if docs_permitidos is not None and nome_sem_ext not in docs_permitidos:
                continue
            caminho = os.path.join(pasta, arquivo)
            arquivos_abertos[nome_sem_ext] = (
                arquivo, open(caminho, 'rb'), 'application/octet-stream'
            )

    try:
        resposta = req_ext.post(
            n8n_url, data=payload_str,
            files=arquivos_abertos if arquivos_abertos else None,
            timeout=60,
        )
        if resposta.status_code >= 400:
            logger.error(f"[N8N] HTTP {resposta.status_code} | UUID: {submission_uuid}")
            return False, f'N8N respondeu HTTP {resposta.status_code}'
        logger.info(f"[N8N] Status: {resposta.status_code} | UUID: {submission_uuid}")
        return True, None
    except Exception as e:
        logger.error(f"[N8N] Erro ao disparar webhook: {e}")
        return False, str(e)
    finally:
        for _nome, (_, fp, _) in arquivos_abertos.items():
            try:
                fp.close()
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════
# RECEBER FICHA DO CLIENTE
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/receber_ficha', methods=['POST'])
def receber_ficha():

    # ── 1. Verificar se é correção ─────────────────────────────
    uuid_existente = request.form.get('uuid', '').strip()
    is_correcao    = bool(uuid_existente)
    link_ficha     = None
    ficha_token    = ''

    if is_correcao:
        sub_original = banco_cadastros.buscar_submissao(uuid_existente)
        if not sub_original:
            return jsonify({'status': 'erro', 'mensagem': 'UUID inválido ou expirado.'}), 400
        submission_uuid = uuid_existente
        docs_originais  = sub_original['docs_enviados']
    else:
        # Ficha nova: exige um token de link válido (V-05), dentro da
        # validade e ainda não utilizado.
        ficha_token = request.form.get('ficha_token', '')
        link_ficha, motivo = banco_links.validar_link(ficha_token)
        if motivo:
            msg = (
                'Este link já foi utilizado para enviar uma ficha.'
                if motivo == 'usado'
                else 'Este link é inválido ou expirou. Solicite um novo link ao seu vendedor.'
            )
            return jsonify({'status': 'erro', 'mensagem': msg}), 403
        submission_uuid = str(uuid_lib.uuid4())
        docs_originais  = []

    # ── 2. Extrair campos de texto ─────────────────────────────
    campos = {
        'vendedor':               request.form.get('vendedor',               ''),
        'representante':          request.form.get('representante',           ''),
        'captacao':               request.form.get('captacao',                ''),
        'tipo_cadastro':          request.form.get('tipo_cadastro',           ''),
        'codigo_nl':              request.form.get('codigo_nl',               ''),
        'contribuinte_icms':      request.form.get('contribuinte_icms',       ''),
        'possui_regime_especial': request.form.get('possui_regime_especial',  ''),
        'regra_faturamento':      request.form.get('regra_faturamento',       ''),
        'cnpj':                   request.form.get('cnpj',                    ''),
        'razao_social':           request.form.get('razao_social',            ''),
        'endereco':               request.form.get('endereco',                ''),
        'telefone_empresa':       request.form.get('telefone_empresa',        ''),
        'whatsapp':               request.form.get('whatsapp',                ''),
        'email_xml_1':            request.form.get('email_xml_1',             ''),
        'email_xml_2':            request.form.get('email_xml_2',             ''),
        'email_xml_3':            request.form.get('email_xml_3',             ''),
        'compras_nome':           request.form.get('compras_nome',            ''),
        'compras_tel':            request.form.get('compras_tel',             ''),
        'compras_email':          request.form.get('compras_email',           ''),
        'rec_nome':               request.form.get('rec_nome',                ''),
        'rec_tel':                request.form.get('rec_tel',                 ''),
        'rec_email':              request.form.get('rec_email',               ''),
        'fin_nome':               request.form.get('fin_nome',                ''),
        'fin_tel':                request.form.get('fin_tel',                 ''),
        'fin_email':              request.form.get('fin_email',               ''),
        'farma_nome':             request.form.get('farma_nome',              ''),
        'farma_tel':              request.form.get('farma_tel',               ''),
        'farma_email':            request.form.get('farma_email',             ''),
    }

    # Em ficha nova, os dados de rastreio vêm do LINK (fonte confiável),
    # não do formulário — assim ninguém adultera vendedor/tipo/etc.
    if not is_correcao and link_ficha:
        campos['vendedor']      = link_ficha.vendedor or campos['vendedor']
        campos['representante'] = link_ficha.representante or campos['representante']
        campos['captacao']      = link_ficha.captacao or campos['captacao']
        campos['tipo_cadastro'] = link_ficha.tipo or campos['tipo_cadastro']
        campos['codigo_nl']     = link_ficha.codigo_nl or campos['codigo_nl']

    # ── 2.1. Enriquecimento via CNPJ.ws ───────────────────────
    cnpj_limpo = re.sub(r'\D', '', campos['cnpj'])

    # O CNPJ preenchido na ficha PRECISA bater com o CNPJ do link.
    # Se não bater, recusa de imediato (não salva, não consome o link)
    # e orienta a pedir um novo link ao vendedor.
    if (not is_correcao and link_ficha and link_ficha.cnpj_cliente
            and link_ficha.cnpj_cliente != cnpj_limpo):
        return jsonify({
            'status': 'erro',
            'mensagem': 'O CNPJ preenchido não corresponde ao CNPJ deste link. '
                        'Verifique o número ou solicite um novo link ao seu vendedor.'
        }), 403

    dados_cnpj_ws = consultar_cnpj_ws(cnpj_limpo)

    campos['cnpj_limpo']               = cnpj_limpo
    campos['nome_fantasia']            = dados_cnpj_ws['nome_fantasia']
    campos['inscricao_estadual']       = dados_cnpj_ws['inscricao_estadual']
    campos['cnae_principal_descricao'] = dados_cnpj_ws['cnae_principal_descricao']
    campos['cep']                      = dados_cnpj_ws['cep']
    campos['logradouro']               = dados_cnpj_ws['logradouro']
    campos['numero']                   = dados_cnpj_ws['numero']
    campos['complemento']              = dados_cnpj_ws['complemento']
    campos['bairro']                   = dados_cnpj_ws['bairro']
    campos['uf']                       = dados_cnpj_ws['uf']
    campos['_cnpj_ws_status']          = dados_cnpj_ws['_cnpj_ws_status']

    # ── 3. Salvar arquivos ─────────────────────────────────────
    lista_docs = [
        'doc_regime_especial', 'doc_alvara', 'doc_crt', 'doc_afe',
        'doc_balanco', 'doc_balancete', 'doc_contrato'
    ]
    pasta_uuid = os.path.join('uploads_fichas', submission_uuid)
    os.makedirs(pasta_uuid, exist_ok=True)

    docs_novos  = []
    docs_manter = json.loads(request.form.get('docs_manter', '[]'))

    for doc_nome in lista_docs:
        arquivo = request.files.get(doc_nome)
        if arquivo and arquivo.filename:
            nome_seguro = secure_filename(arquivo.filename)
            ext         = os.path.splitext(nome_seguro)[1].lower() or '.pdf'
            caminho     = os.path.join(pasta_uuid, f"{doc_nome}{ext}")
            arquivo.save(caminho)
            docs_novos.append(doc_nome)

    # ── 4. Calcular docs finais ────────────────────────────────
    docs_mantidos_validos = [d for d in docs_manter if d in docs_originais]
    docs_finais           = list(set(docs_novos + docs_mantidos_validos))

    # ── 5. Categorizar documentos ──────────────────────────────
    if is_correcao:
        docs_ja_existentes = [d for d in docs_mantidos_validos]
        docs_substituidos  = [d for d in docs_novos if d in docs_originais]
        docs_adicionados   = [d for d in docs_novos if d not in docs_originais]
        nova_tentativa = banco_cadastros.incrementar_tentativa(
            submission_uuid, campos, docs_finais
        )
    else:
        docs_ja_existentes = []
        docs_substituidos  = []
        docs_adicionados   = []
        nova_tentativa     = 1
        banco_cadastros.salvar_submissao(
            submission_uuid, campos['cnpj'], campos['razao_social'],
            campos, docs_finais
        )
        # Marca o link como usado (o CNPJ já foi validado contra o do link).
        if link_ficha:
            banco_links.marcar_usado(ficha_token, submission_uuid, False)

    # ── 6. Nomes legíveis dos docs ─────────────────────────────
    nomes_legiveis = {
        'doc_regime_especial': 'Regime Especial de ICMS',
        'doc_alvara':          'Alvará Sanitário',
        'doc_crt':             'Certidão de Regularidade',
        'doc_afe':             'Publicação AFE / AE',
        'doc_balanco':         'Último Balanço e DRE',
        'doc_balancete':       'Balancete (< 90 dias)',
        'doc_contrato':        'Contrato Social',
    }

    def nomes(lista):
        return ', '.join(nomes_legiveis.get(d, d) for d in lista) if lista else '—'

    # ── 7. Payload para o N8N ──────────────────────────────────
    payload_n8n = {
        **campos,
        'uuid':               submission_uuid,
        'tentativa':          nova_tentativa,
        'is_correcao':        str(is_correcao).lower(),
        'docs_ja_existentes': nomes(docs_ja_existentes),
        'docs_substituidos':  nomes(docs_substituidos),
        'docs_adicionados':   nomes(docs_adicionados),
        'qtd_documentos':     len(docs_finais),
    }

    # ── 8. Disparar para o N8N (com controle anti-ficha-órfã) ──
    payload_str = {k: str(v) for k, v in payload_n8n.items()}
    # Guarda o payload para permitir reenvio pelo Monitor se o N8N falhar.
    banco_cadastros.salvar_payload_n8n(submission_uuid, payload_str)

    ok, erro = _disparar_para_n8n(submission_uuid, payload_str, docs_finais)
    if ok:
        banco_cadastros.marcar_envio_n8n(submission_uuid, 'enviado')
    else:
        # A ficha JÁ está salva no banco. Marcamos o envio como falho para o
        # operador reenviar pelo Monitor — assim a ficha não fica "órfã".
        banco_cadastros.marcar_envio_n8n(submission_uuid, 'falhou', erro)

    # Ficha sem documentos anexados: o n8n (via nó IF) vai avisar o vendedor
    # pra reenviar. Marcamos como "aguardando_documentos" pra dar visibilidade
    # no Monitor. (salvar_submissao/incrementar_tentativa já deixaram 'pendente';
    # aqui sobrescrevemos apenas quando faltam documentos.)
    if not docs_finais:
        banco_cadastros.atualizar_status(submission_uuid, 'aguardando_documentos')

    # Para o vendedor, a ficha foi recebida nos dois casos (ela está salva).
    # Se o N8N estiver fora do ar, a equipe reenvia internamente pelo Monitor.
    return jsonify({'status': 'ok', 'uuid': submission_uuid})


# ══════════════════════════════════════════════════════════════
# API: dados de uma submissão (pré-preenchimento na correção)
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/api/ficha/<submission_uuid>')
def api_ficha(submission_uuid):
    sub = banco_cadastros.buscar_submissao(submission_uuid)
    if not sub:
        return jsonify({'erro': 'Submissão não encontrada'}), 404
    return jsonify({
        'dados':              sub['dados_json'],
        'docs_enviados':      sub['docs_enviados'],
        'motivos_reprovacao': sub['motivos_reprovacao'],
        'tentativa':          sub['tentativa'],
        'razao_social':       sub['razao_social'],
        'cnpj':               sub['cnpj'],
    })


# ══════════════════════════════════════════════════════════════
# REENVIAR FICHA AO N8N (anti-ficha-órfã — operador no Monitor)
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/api/reenviar_ficha/<uuid>', methods=['POST'])
@permissao_required('monitor')
def reenviar_ficha(uuid):
    """
    Reenvia ao N8N uma ficha cujo envio falhou. Usa o payload guardado e
    reanexa os documentos da pasta. Atualiza o status do envio.
    """
    if not eh_uuid_valido(uuid):
        return jsonify({'erro': 'UUID inválido.'}), 400

    sub = banco_cadastros.buscar_submissao(uuid)
    if not sub:
        return jsonify({'erro': 'Submissão não encontrada.'}), 404

    payload = banco_cadastros.buscar_payload_n8n(uuid)
    if not payload:
        return jsonify({'erro': 'Não há dados de envio guardados para esta ficha.'}), 400

    ok, erro = _disparar_para_n8n(uuid, payload, sub.get('docs_enviados'))
    if ok:
        banco_cadastros.marcar_envio_n8n(uuid, 'enviado')
        return jsonify({'status': 'ok'})

    banco_cadastros.marcar_envio_n8n(uuid, 'falhou', erro)
    return jsonify({'status': 'erro', 'mensagem': erro or 'Falha ao reenviar.'}), 502


# ══════════════════════════════════════════════════════════════
# MONITOR DE LINKS GERADOS (controle/rastreio)
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/monitor_links')
@permissao_required('monitor')
def monitor_links():
    return render_template('monitor_links.html')


@clientes_bp.route('/api/monitor_links')
@permissao_required('monitor')
def api_monitor_links():
    links = banco_links.listar_links(
        busca    = request.args.get('busca'),
        status   = request.args.get('status'),
        data_de  = request.args.get('data_de'),
        data_ate = request.args.get('data_ate'),
    )
    stats = banco_links.estatisticas()
    return jsonify({'links': links, 'stats': stats})


@clientes_bp.route('/api/links/<int:link_id>', methods=['DELETE'])
@admin_required
def excluir_link(link_id):
    """Exclui um link gerado. Apenas administradores do Hub (admin_required)."""
    if banco_links.excluir_link(link_id):
        return jsonify({'status': 'ok'})
    return jsonify({'erro': 'Link não encontrado.'}), 404


# ══════════════════════════════════════════════════════════════
# Aprovação / Reprovação (farmacêutica via email)
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/confirmar')
def confirmar_acao():
    acao           = request.args.get('acao',           '')
    empresa        = request.args.get('empresa',        '')
    resume_url     = request.args.get('resumeUrl',      '')
    motivo         = request.args.get('motivo',         '')
    uuid_ficha     = request.args.get('uuid',           '')
    acao_bloqueio  = request.args.get('acao_bloqueio',  '')
    valor_bloqueio = request.args.get('valor_bloqueio', '')

    if not resume_url:
        return "Parâmetro resumeUrl ausente.", 400

    # Anti-SSRF (V-04): só seguimos se a resume_url for do N8N oficial.
    if not resume_url_confiavel(resume_url):
        return "Parâmetro resumeUrl inválido.", 400

    try:
        if acao == 'aprovar':
            # Monta params pro n8n. Sempre manda decisao=aprovar.
            # Se vier acao_bloqueio (de uma reativação que passou pela
            # tela decidir_bloqueio), propaga junto. Cadastros novos
            # não mandam esses params — eles chegam vazios e o n8n ignora.
            params_n8n = {'decisao': 'aprovar'}
            if acao_bloqueio:
                params_n8n['acao_bloqueio']  = acao_bloqueio
            if valor_bloqueio:
                params_n8n['valor_bloqueio'] = valor_bloqueio

            req_ext.get(resume_url, params=params_n8n, timeout=10)

            if uuid_ficha:
                banco_cadastros.atualizar_status(uuid_ficha, 'aprovado')

        elif acao == 'reprovar':
            req_ext.get(resume_url, params={'decisao': 'reprovar', 'motivo': motivo}, timeout=10)
            if uuid_ficha and motivo:
                banco_cadastros.registrar_reprovacao(uuid_ficha, motivo)

        elif acao == 'mascaras':
            req_ext.get(resume_url, params={'mascaras': 'confirmadas'}, timeout=10)

    except Exception as e:
        logger.warning(f"Aviso ao chamar N8N: {e}")

    return render_template('confirmar_acao.html', acao=acao, empresa=empresa, motivo=motivo)


@clientes_bp.route('/motivo_reprovacao')
def motivo_reprovacao():
    resume_url = request.args.get('resumeUrl', '')
    empresa    = request.args.get('empresa',   '')
    cnpj       = request.args.get('cnpj',      '')
    uuid_ficha = request.args.get('uuid',      '')
    return render_template(
        'motivo_reprovacao.html',
        resume_url=resume_url, empresa=empresa,
        cnpj=cnpj, uuid=uuid_ficha
    )

@clientes_bp.route('/decidir_bloqueio')
def decidir_bloqueio():
    """
    Tela intermediária para a farmacêutica decidir o bloqueio
    ao aprovar uma REATIVAÇÃO. Só é chamada para reativações —
    cadastros novos vão direto pro /confirmar.

    Parâmetros recebidos via query string (vindos do link do email):
      resumeUrl → pra propagar pro /confirmar depois
      empresa   → nome da empresa (exibição)
      uuid      → UUID da submissão
    """
    resume_url = request.args.get('resumeUrl', '')
    empresa    = request.args.get('empresa',   '')
    uuid_ficha = request.args.get('uuid',      '')

    if not resume_url:
        return "Parâmetro resumeUrl ausente.", 400

    return render_template(
        'decidir_bloqueio.html',
        resume_url=resume_url,
        empresa=empresa,
        uuid=uuid_ficha
    )


# ══════════════════════════════════════════════════════════════
# Monitor de Cadastros
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/monitor_cadastros')
@permissao_required('monitor')
def monitor_cadastros():
    return render_template('monitor_cadastros.html')


@clientes_bp.route('/api/monitor_cadastros')
@permissao_required('monitor')
def api_monitor_cadastros():
    cadastros = banco_cadastros.listar_submissoes(
        status   = request.args.get('status'),
        tipo     = request.args.get('tipo'),
        busca    = request.args.get('busca'),
        data_de  = request.args.get('data_de'),
        data_ate = request.args.get('data_ate'),
    )
    stats = banco_cadastros.stats_submissoes()
    return jsonify({'cadastros': cadastros, 'stats': stats})


@clientes_bp.route('/download_doc/<uuid>/<doc_nome>')
@permissao_required('monitor')
def download_doc(uuid, doc_nome):
    # O 'uuid' vem da URL e é usado para montar um caminho de pasta.
    # Sem validar, alguém poderia passar algo como "../../etc" e escapar
    # da pasta de uploads (path traversal). Exigimos um UUID bem-formado.
    if not eh_uuid_valido(uuid):
        return "❌ Documento não encontrado.", 404

    base = os.path.join(os.getcwd(), 'uploads_fichas')
    pasta_cliente = os.path.join(base, uuid)
    # Defesa extra: garante que o caminho final está realmente dentro da
    # pasta de uploads, mesmo que algo escape da validação acima.
    if os.path.commonpath([os.path.realpath(pasta_cliente), os.path.realpath(base)]) != os.path.realpath(base):
        return "❌ Documento não encontrado.", 404

    if os.path.isdir(pasta_cliente):
        for arquivo in os.listdir(pasta_cliente):
            nome_sem_extensao = os.path.splitext(arquivo)[0]
            if nome_sem_extensao == doc_nome:
                caminho_completo = os.path.join(pasta_cliente, arquivo)
                return send_file(caminho_completo, as_attachment=False)
    return "❌ Documento não encontrado.", 404


# ══════════════════════════════════════════════════════════════
# Exportar fichas como Excel
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/api/exportar_fichas')
@permissao_required('monitor')
def exportar_fichas():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    registros = banco_cadastros.listar_submissoes(
        status   = request.args.get('status'),
        tipo     = request.args.get('tipo'),
        busca    = request.args.get('busca'),
        data_de  = request.args.get('data_de'),
        data_ate = request.args.get('data_ate'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Fichas Cadastrais"

    COR_TITULO_BG = "1A2744"
    COR_HEADER_BG = "0EE6B8"
    COR_HEADER_FT = "090D14"
    COR_LINHA_PAR = "EEF4FF"
    COR_LINHA_IMP = "FFFFFF"
    COR_STATUS = {
        "pendente":   ("FEF3C7", "92400E"),
        "aprovado":   ("D1FAE5", "065F46"),
        "reprovado":  ("FEE2E2", "991B1B"),
        "cadastrado": ("EDE9FE", "5B21B6"),
    }

    def fill(cor):
        return PatternFill("solid", start_color=cor, end_color=cor)

    def borda():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def cel_fmt(cell, bold=False, cor_ft="000000", cor_bg=None, size=9, wrap=False, h_align="left"):
        cell.font      = Font(name="Arial", bold=bold, color=cor_ft, size=size)
        cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
        cell.border    = borda()
        if cor_bg:
            cell.fill = fill(cor_bg)

    # Título
    ws.merge_cells("A1:R1")
    ws["A1"] = "CIAMED  —  Relatório de Fichas Cadastrais"
    ws["A1"].font      = Font(name="Arial", bold=True, size=15, color=COR_TITULO_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill      = fill("F0F4FF")
    ws.row_dimensions[1].height = 30

    # Subtítulo
    params = request.args
    partes = []
    if params.get("status"):   partes.append(f"Status: {params['status']}")
    if params.get("tipo"):     partes.append(f"Tipo: {params['tipo']}")
    if params.get("busca"):    partes.append(f"Busca: {params['busca']}")
    if params.get("data_de"):  partes.append(f"De: {params['data_de']}")
    if params.get("data_ate"): partes.append(f"Até: {params['data_ate']}")
    filtros_str = "  |  ".join(partes) if partes else "Sem filtros"

    ws.merge_cells("A2:R2")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}     •     Filtros: {filtros_str}     •     Total: {len(registros)} ficha(s)"
    ws["A2"].font      = Font(name="Arial", size=9, color="555555", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].fill      = fill("F0F4FF")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Cabeçalhos
    COLUNAS = [
        ("Data/Hora", 18), ("Razão Social", 32), ("CNPJ", 17),
        ("Tipo", 14), ("Status", 14), ("Tentativa", 9),
        ("Cód. ERP", 12), ("Vendedor", 15), ("Representante", 16),
        ("Captação", 15), ("ICMS Contribuinte", 15), ("Regime Especial", 15),
        ("Regra Faturamento", 15), ("Telefone", 14), ("WhatsApp", 14),
        ("E-mail XML", 28), ("Documentos", 38), ("Motivo Reprovação", 42),
    ]

    NOMES_DOCS = {
        "doc_regime_especial": "Regime Especial ICMS",
        "doc_alvara": "Alvará Sanitário",
        "doc_crt": "Certidão de Regularidade",
        "doc_afe": "AFE / AE",
        "doc_balanco": "Balanço e DRE",
        "doc_balancete": "Balancete",
        "doc_contrato": "Contrato Social",
    }

    LIN_HEADER = 4
    for col_idx, (titulo, largura) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=LIN_HEADER, column=col_idx, value=titulo)
        cel_fmt(cell, bold=True, cor_ft=COR_HEADER_FT, cor_bg=COR_HEADER_BG, size=9, h_align="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[LIN_HEADER].height = 22

    # Dados
    LABEL_STATUS = {
        "pendente": "⏳ Em Análise", "aprovado": "✅ Aprovado",
        "reprovado": "❌ Reprovado", "cadastrado": "🤖 Cadastrado ERP",
    }

    for i, c in enumerate(registros):
        lin    = LIN_HEADER + 1 + i
        d      = c.get("dados_json") or {}
        status = c.get("status", "pendente")
        bg_lin = COR_LINHA_PAR if i % 2 == 0 else COR_LINHA_IMP

        docs_lista = [NOMES_DOCS.get(x, x) for x in (c.get("docs_enviados") or [])]
        docs_str   = ", ".join(docs_lista) if docs_lista else "—"
        motivos    = c.get("motivos_reprovacao") or []
        motivo_str = motivos[-1]["motivo"] if motivos else "—"

        valores = [
            c.get("created_at", "—"), c.get("razao_social", "—"), c.get("cnpj", "—"),
            d.get("tipo_cadastro", "—"), LABEL_STATUS.get(status, status),
            f"{c.get('tentativa', 1)}ª", d.get("codigo_nl", "—"),
            d.get("vendedor", "—"), d.get("representante", "—"),
            d.get("captacao", "—"), d.get("contribuinte_icms", "—"),
            d.get("possui_regime_especial", "—"), d.get("regra_faturamento", "—"),
            d.get("telefone_empresa", "—"), d.get("whatsapp", "—"),
            d.get("email_xml_1", "—"), docs_str, motivo_str,
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=lin, column=col_idx, value=valor)
            h_al = "center" if col_idx in [4, 5, 6, 7] else "left"
            if col_idx == 5:
                bg_s, ft_s = COR_STATUS.get(status, (bg_lin, "000000"))
                cel_fmt(cell, cor_ft=ft_s, cor_bg=bg_s, size=9, h_align="center", bold=True)
            else:
                cel_fmt(cell, cor_ft="222222", cor_bg=bg_lin, size=9, h_align=h_al, wrap=(col_idx in [16, 17, 18]))
        ws.row_dimensions[lin].height = 18

    ws.freeze_panes = "A5"
    ultima_col = get_column_letter(len(COLUNAS))
    ws.auto_filter.ref = f"A{LIN_HEADER}:{ultima_col}{LIN_HEADER + len(registros)}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nome_arquivo = f"Fichas_Cadastrais_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=nome_arquivo,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════════
# Fila de Cadastro (integração N8N)
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/fila_cadastro/entrar', methods=['POST'])
@exigir_token_n8n
def fila_cadastro_entrar():
    dados      = request.get_json()
    resume_url = dados.get('resume_url', '')
    if not resume_url:
        return jsonify({'erro': 'resume_url obrigatório'}), 400

    # Anti-SSRF (V-04): a URL é guardada e depois "visitada" pelo Hub.
    if not resume_url_confiavel(resume_url):
        return jsonify({'erro': 'resume_url inválida'}), 400

    novo_id, posicao, sua_vez = cadastro_entrar(resume_url)
    return jsonify({'id': novo_id, 'posicao': posicao, 'sua_vez': sua_vez})


@clientes_bp.route('/fila_cadastro/liberar_proximo', methods=['POST'])
@exigir_token_n8n
def fila_cadastro_liberar():
    cadastro_liberar_proximo()
    return jsonify({'status': 'ok'})

# ══════════════════════════════════════════════════════════════
# ADICIONAR no routes/clientes.py — LOGO ABAIXO da rota
# /fila_cadastro/liberar_proximo
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/fila_cadastro/status', methods=['GET'])
def fila_cadastro_status():
    """Mostra o estado atual da fila de cadastro — útil pra debug."""
    return jsonify(fila_cadastro_status_dict())


@clientes_bp.route('/fila_cadastro/reset', methods=['POST'])
def fila_cadastro_reset():
    """
    Destrava a fila forçadamente. Usa quando:
      - O robô falhou e a fila ficou presa
      - Tem execuções fantasma que nunca vão terminar

    O que faz:
      1. Remove TODAS as linhas da fila de cadastro no banco
         (pendentes e a que estava executando)
      2. NÃO acorda nenhum Wait node do N8N (as execuções
         pendentes no N8N vão expirar sozinhas pelo timeout)
    """
    qtd_removidos = _fila_cadastro_reset()

    return jsonify({
        "status": "ok",
        "mensagem": f"Fila resetada. {qtd_removidos} item(ns) removido(s).",
        "em_execucao": False,
        "tamanho_fila": 0
    })


# ══════════════════════════════════════════════════════════════
# PEÇA 3A — Rota /n8n_iniciar_cadastro ATUALIZADA
# ══════════════════════════════════════════════════════════════
# Substitui a versão anterior (Entrega 1).
# Diferença: agora trata o retorno padronizado do robô e
# registra erro_robo no banco quando falha.
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/n8n_iniciar_cadastro', methods=['POST'])
@exigir_token_n8n
def n8n_iniciar_cadastro():
    dados_n8n = request.json or {}

    # ── 1. Validar UUID obrigatório ──
    uuid_ficha = dados_n8n.get('uuid', '').strip()
    if not uuid_ficha:
        return jsonify({"status": "erro", "detalhes": "Campo 'uuid' ausente no payload."}), 400

    # ── 2. Buscar dados completos no banco ──
    sub = banco_cadastros.buscar_submissao(uuid_ficha)
    if not sub:
        return jsonify({"status": "erro", "detalhes": f"UUID {uuid_ficha} não encontrado no banco."}), 404

    d = sub['dados_json']

    if d.get('tipo_cadastro') != 'NOVO':
        return jsonify({
            "status": "erro",
            "detalhes": f"Submissão {uuid_ficha} não é NOVO (tipo: {d.get('tipo_cadastro')})."
        }), 400

    razao = d.get('razao_social', '') or sub.get('razao_social', '')
    logger.info(f"[N8N] Cadastro NOVO recebido | Empresa: {razao} | UUID: {uuid_ficha}")

    # ── 3. Montar contatos ──
    contatos = []
    if d.get('compras_nome'):
        contatos.append({
            "nome": d['compras_nome'], "codigo": "30",
            "tel": d.get('compras_tel', ''), "email": d.get('compras_email', ''),
            "email_xml": d.get('email_xml_1', ''), "email_danfe": "",
            "check_boleto": False, "check_docs": True,
        })
    if d.get('rec_nome'):
        contatos.append({
            "nome": d['rec_nome'], "codigo": "35",
            "tel": d.get('rec_tel', ''), "email": d.get('rec_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": False, "check_docs": False,
        })
    if d.get('fin_nome'):
        contatos.append({
            "nome": d['fin_nome'], "codigo": "50",
            "tel": d.get('fin_tel', ''), "email": d.get('fin_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": True, "check_docs": False,
        })
    if d.get('farma_nome'):
        contatos.append({
            "nome": d['farma_nome'], "codigo": "20",
            "tel": d.get('farma_tel', ''), "email": d.get('farma_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": False, "check_docs": False,
        })

    # ── 4. Mapear pro formato do robô ──
    dados_robo = {
        "cnpj_limpo":               d.get('cnpj_limpo', ''),
        "razao_social":             razao,
        "nome_fantasia":            d.get('nome_fantasia', ''),
        "inscricao_estadual":       d.get('inscricao_estadual', ''),
        "cnae_principal_descricao": d.get('cnae_principal_descricao', ''),
        "cep":                      d.get('cep', ''),
        "logradouro":               d.get('logradouro', ''),
        "numero":                   d.get('numero', ''),
        "complemento":              d.get('complemento', ''),
        "bairro":                   d.get('bairro', ''),
        "uf":                       d.get('uf', ''),
        "regra_faturamento":        d.get('regra_faturamento', ''),
        "representante":            d.get('representante', ''),
        "vendedor":                 d.get('vendedor', ''),
        "forma_captacao":           d.get('captacao', ''),
        "email_xml":                d.get('email_xml_1', ''),
        "contatos_da_tela":         contatos,
        "telefone_empresa":         d.get('telefone_empresa', ''),
        "tem_cebas":                consultar_cebas(d.get('cnpj_limpo', '')),
    }

    # ── 5. Disparar o robô (com tratamento de erro completo) ──
    exec_id = iniciar_execucao_robo('cliente_novo')
    try:
        resultado = cadastro_novo.executar(dados_robo)

        # O robô SEMPRE retorna um dict com "status"
        status_robo = resultado.get('status', 'Erro')

        if status_robo in ('Sucesso', 'Sucesso com avisos'):
            finalizar_execucao_robo(exec_id, 'sucesso', resultado)
            banco_cadastros.atualizar_status(uuid_ficha, 'cadastrado')

            return jsonify({
                "status":         "sucesso",
                "mensagem":       f"Robô finalizou o cadastro de {razao}",
                "resultado":      resultado,
                "codigo_cliente": resultado.get('codigo_cliente'),
            }), 200

        else:
            # Robô retornou status de erro (controlado, não exceção)
            finalizar_execucao_robo(exec_id, 'erro', resultado)
            banco_cadastros.registrar_erro_robo(uuid_ficha, {
                "etapa_falha":     resultado.get('etapa_falha'),
                "etapa_concluida": resultado.get('etapa_concluida'),
                "erro_detalhe":    resultado.get('erro_detalhe', 'Erro não especificado'),
                "data_erro":       datetime.now().strftime('%d/%m/%Y %H:%M'),
                "tentativa_robo":  sub.get('tentativa', 1),
            })

            return jsonify({
                "status":    "erro_robo",
                "mensagem":  f"Robô falhou no cadastro de {razao}",
                "resultado": resultado,
            }), 200  # ← 200 propositalmente! O N8N precisa receber o body pra decidir no Switch

    except Exception as e:
        # Exceção não tratada (crash do robô)
        finalizar_execucao_robo(exec_id, 'erro', {'erro': str(e)})
        banco_cadastros.registrar_erro_robo(uuid_ficha, {
            "etapa_falha":     "excecao_nao_tratada",
            "etapa_concluida": None,
            "erro_detalhe":    str(e),
            "data_erro":       datetime.now().strftime('%d/%m/%Y %H:%M'),
            "tentativa_robo":  sub.get('tentativa', 1),
        })
        logger.error(f"[N8N] Exceção no cadastro NOVO de {razao}: {e}")

        return jsonify({
            "status":   "erro_robo",
            "mensagem": f"Exceção no cadastro de {razao}: {str(e)}",
        }), 200  # ← 200 pra o N8N conseguir ler o body


# ══════════════════════════════════════════════════════════════
# PEÇA 3B — Rota /n8n_iniciar_reativacao ATUALIZADA
# ══════════════════════════════════════════════════════════════
# Mesma lógica de tratamento de erro da rota de cadastro novo.
# Substitui a versão atual no routes/clientes.py.
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/n8n_iniciar_reativacao', methods=['POST'])
@exigir_token_n8n
def n8n_iniciar_reativacao():
    dados_n8n = request.json or {}

    # ── 1. Validar campos obrigatórios ──
    uuid_ficha = dados_n8n.get('uuid', '').strip()
    if not uuid_ficha:
        return jsonify({"status": "erro", "detalhes": "Campo 'uuid' ausente no payload."}), 400

    # ── 2. Buscar dados completos no banco ──
    sub = banco_cadastros.buscar_submissao(uuid_ficha)
    if not sub:
        return jsonify({"status": "erro", "detalhes": f"UUID {uuid_ficha} não encontrado no banco."}), 404

    d = sub['dados_json']

    if d.get('tipo_cadastro') != 'REATIVACAO':
        return jsonify({
            "status": "erro",
            "detalhes": f"Submissão {uuid_ficha} não é REATIVACAO (tipo: {d.get('tipo_cadastro')})."
        }), 400

    codigo_nl = d.get('codigo_nl', '').strip()
    if not codigo_nl:
        return jsonify({"status": "erro", "detalhes": "Código NL ausente nos dados da submissão."}), 400

    razao = d.get('razao_social', '') or sub.get('razao_social', '')
    logger.info(f"[N8N] Reativação recebida | NL: {codigo_nl} | Empresa: {razao}")

    # ── 3. Montar contatos ──
    contatos = []
    if d.get('compras_nome'):
        contatos.append({
            "nome": d['compras_nome'], "codigo": "30",
            "tel": d.get('compras_tel', ''), "email": d.get('compras_email', ''),
            "email_xml": d.get('email_xml_1', ''), "email_danfe": "",
            "check_boleto": False, "check_docs": True,
        })
    if d.get('rec_nome'):
        contatos.append({
            "nome": d['rec_nome'], "codigo": "35",
            "tel": d.get('rec_tel', ''), "email": d.get('rec_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": False, "check_docs": False,
        })
    if d.get('fin_nome'):
        contatos.append({
            "nome": d['fin_nome'], "codigo": "50",
            "tel": d.get('fin_tel', ''), "email": d.get('fin_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": True, "check_docs": False,
        })
    if d.get('farma_nome'):
        contatos.append({
            "nome": d['farma_nome'], "codigo": "20",
            "tel": d.get('farma_tel', ''), "email": d.get('farma_email', ''),
            "email_xml": "", "email_danfe": "",
            "check_boleto": False, "check_docs": False,
        })

    # ── 4. Mapear pro formato do robô ──
    dados_robo = {
        "codigo_nl":                codigo_nl,
        "cnpj_limpo":               d.get('cnpj_limpo', ''),
        "razao_social":             razao,
        "nome_fantasia":            d.get('nome_fantasia', ''),
        "inscricao_estadual":       d.get('inscricao_estadual', ''),
        "cnae_principal_descricao": d.get('cnae_principal_descricao', ''),
        "cep":                      d.get('cep', ''),
        "logradouro":               d.get('logradouro', ''),
        "numero":                   d.get('numero', ''),
        "complemento":              d.get('complemento', ''),
        "bairro":                   d.get('bairro', ''),
        "uf":                       d.get('uf', ''),
        "acao_bloqueio":            dados_n8n.get('acao_bloqueio'),
        "valor_bloqueio":           dados_n8n.get('valor_bloqueio'),
        "regra_faturamento":        d.get('regra_faturamento', ''),
        "representante":            d.get('representante', ''),
        "vendedor":                 d.get('vendedor', ''),
        "forma_captacao":           d.get('captacao', ''),
        "email_xml":                d.get('email_xml_1', ''),
        "contatos_da_tela":         contatos,
        "telefone_empresa":         d.get('telefone_empresa', ''),
        "tem_cebas":                consultar_cebas(d.get('cnpj_limpo', '')),
    }

    # ── 5. Disparar o robô (com tratamento de erro completo) ──
    exec_id = iniciar_execucao_robo('cliente_reativacao')
    try:
        resultado = cadastro_reativacao.executar(dados_robo)
        status_robo = resultado.get('status', 'Erro')

        if status_robo in ('Sucesso', 'Sucesso com avisos'):
            finalizar_execucao_robo(exec_id, 'sucesso', resultado)
            banco_cadastros.atualizar_status(uuid_ficha, 'cadastrado')

            return jsonify({
                "status":         "sucesso",
                "mensagem":       f"Robô finalizou a reativação do NL {codigo_nl}",
                "resultado":      resultado,
                "codigo_cliente": resultado.get('codigo_cliente'),
            }), 200

        else:
            finalizar_execucao_robo(exec_id, 'erro', resultado)
            banco_cadastros.registrar_erro_robo(uuid_ficha, {
                "etapa_falha":     resultado.get('etapa_falha'),
                "etapa_concluida": resultado.get('etapa_concluida'),
                "erro_detalhe":    resultado.get('erro_detalhe', 'Erro não especificado'),
                "data_erro":       datetime.now().strftime('%d/%m/%Y %H:%M'),
                "tentativa_robo":  sub.get('tentativa', 1),
            })

            return jsonify({
                "status":    "erro_robo",
                "mensagem":  f"Robô falhou na reativação do NL {codigo_nl}",
                "resultado": resultado,
            }), 200

    except Exception as e:
        finalizar_execucao_robo(exec_id, 'erro', {'erro': str(e)})
        banco_cadastros.registrar_erro_robo(uuid_ficha, {
            "etapa_falha":     "excecao_nao_tratada",
            "etapa_concluida": None,
            "erro_detalhe":    str(e),
            "data_erro":       datetime.now().strftime('%d/%m/%Y %H:%M'),
            "tentativa_robo":  sub.get('tentativa', 1),
        })
        logger.error(f"[N8N] Exceção na reativação do NL {codigo_nl}: {e}")

        return jsonify({
            "status":   "erro_robo",
            "mensagem": f"Exceção na reativação de {razao}: {str(e)}",
        }), 200


# ══════════════════════════════════════════════════════════════
# PEÇA 3C — Rota /reprocessar_cadastro (NOVA)
# ══════════════════════════════════════════════════════════════
# Botão do monitor de cadastros chama esta rota.
# Prepara a submissão pra re-entrar na fila sem re-aprovação.
# ══════════════════════════════════════════════════════════════

@clientes_bp.route('/reprocessar_cadastro', methods=['POST'])
@permissao_required('clientes')
def reprocessar_cadastro():
    """
    Reprocessa uma submissão que falhou no robô.
    Muda o status de 'erro_robo' → 'aprovado' e dispara o N8N.

    O N8N vai ver o status 'aprovado' e colocar na fila normalmente,
    como se fosse a primeira vez — mas sem passar pela aprovação
    da farmacêutica de novo.
    """
    dados = request.json or {}
    uuid_ficha = dados.get('uuid', '').strip()

    if not uuid_ficha:
        return jsonify({"status": "erro", "mensagem": "UUID não informado."}), 400

    # Tenta reprocessar no banco
    sucesso = banco_cadastros.reprocessar_submissao(uuid_ficha)
    if not sucesso:
        return jsonify({
            "status": "erro",
            "mensagem": "Não foi possível reprocessar. Verifique se o status é 'erro_robo'."
        }), 400

    # Busca dados pra saber o tipo (NOVO ou REATIVACAO)
    sub = banco_cadastros.buscar_submissao(uuid_ficha)
    d = sub['dados_json']
    tipo = d.get('tipo_cadastro', 'NOVO')

    logger.info(f"[REPROCESSAR] UUID {uuid_ficha} preparado | Tipo: {tipo}")

    # ── Disparar o webhook do N8N pra re-entrar na fila ──
    # O N8N recebe o UUID e coloca na fila como se fosse novo.
    # A farmacêutica NÃO é acionada — vai direto pro robô.
    # -- Disparar o webhook do N8N pra re-entrar na fila --
    N8N_WEBHOOK_URL = os.getenv('N8N_WEBHOOK_REPROCESSAR')

    # Preparar o payload convertendo todos os valores para string (idêntico ao receber_ficha)
    payload_n8n = {
        **d,  # Desempacota os dados originais que vieram do banco_cadastros
        'uuid': uuid_ficha,
        'tipo_cadastro': tipo
    }
    payload_str = {k: str(v) for k, v in payload_n8n.items()}

    arquivos_abertos = {}
    pasta = os.path.join('uploads_fichas', uuid_ficha)

    # Identificar e abrir os arquivos físicos salvos localmente para enviar ao N8N
    if os.path.exists(pasta):
        for arquivo in os.listdir(pasta):
            nome_sem_ext = os.path.splitext(arquivo)[0]
            caminho = os.path.join(pasta, arquivo)
            
            arquivos_abertos[nome_sem_ext] = (
                arquivo, 
                open(caminho, 'rb'), 
                'application/octet-stream'
            )

    try:
        resposta_n8n = req_ext.post(
            N8N_WEBHOOK_URL, 
            data=payload_str,
            files=arquivos_abertos if arquivos_abertos else None,
            timeout=60
        )
        
        logger.info(f"[REPROCESSAR] Status N8N: {resposta_n8n.status_code} | UUID: {uuid_ficha}")

    except Exception as e:
        logger.error(f"[REPROCESSAR] Erro ao disparar webhook: {e}")
        return jsonify({
            'status': 'erro', 
            'mensagem': f'Erro ao disparar reprocessamento no N8N: {str(e)}'
        }), 500
        
    finally:
        # Garante o fechamento de todos os arquivos abertos para evitar vazamento de memória
        for nome, (_, fp, _) in arquivos_abertos.items():
            try:
                fp.close()
            except Exception:
                pass

    return jsonify({
        "status":   "ok",
        "mensagem": f"Submissão {uuid_ficha[:8]}... preparada para reprocessamento.",
        "tipo":     tipo,
    })